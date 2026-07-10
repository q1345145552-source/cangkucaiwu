from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime
from app.database import get_db
from app.models.recharge import IncomingFlow, CurrencyEnum, MatchStatus
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, get_wh_ids, Role, check_staff_permission
from app.schemas.business import IncomingCreate, IncomingBatchImport

router = APIRouter()

@router.get("")
async def list_incoming(
    page: int = 1, page_size: int = 20, month: str = None, status: str = None,
    current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role == Role.STAFF and "到账流水" not in (current_user.extra_permissions or []):
        raise HTTPException(403, "无录入到账流水权限")
    query = select(IncomingFlow); count_q = select(func.count(IncomingFlow.id))
    query = query.where(IncomingFlow.warehouse_id.in_(get_wh_ids(current_user)))
    count_q = count_q.where(IncomingFlow.warehouse_id.in_(get_wh_ids(current_user)))
    if month:
        query = query.where(func.to_char(IncomingFlow.received_date, 'YYYY-MM') == month)
        count_q = count_q.where(func.to_char(IncomingFlow.received_date, 'YYYY-MM') == month)
    if status:
        query = query.where(IncomingFlow.match_status == status)
        count_q = count_q.where(IncomingFlow.match_status == status)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(IncomingFlow.created_at.desc()).offset((page-1)*page_size).limit(page_size))
    records = result.scalars().all()
    user_ids = list(set(r.entrant_id for r in records if r.entrant_id))
    user_map = {}
    if user_ids:
        users = (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all()
        user_map = {u.id: u.display_name for u in users}
    return {"data": [{
        "id": r.id, "warehouse_id": r.warehouse_id,
        "received_date": r.received_date.isoformat() if r.received_date else None,
        "amount": r.amount, "currency": r.currency or "THB",
        "payer_name": r.payer_name, "payment_method": r.payment_method, "remark": r.remark,
        "match_status": r.match_status or "unmatched",
        "entrant_name": user_map.get(r.entrant_id, ""),
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in records], "total": total, "page": page, "page_size": page_size}

@router.post("")
async def create_incoming(req: IncomingCreate, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "仅仓库管理员可操作")
    if current_user.role == Role.STAFF and "到账流水" not in (current_user.extra_permissions or []):
        raise HTTPException(403, "无录入到账流水权限")
    wh_id = get_wh_id(current_user)
    if not wh_id:
        raise HTTPException(400, "无法确定仓库")
    r = IncomingFlow(
        warehouse_id=wh_id,
        received_date=datetime.fromisoformat(req.received_date),
        amount=req.amount, currency=req.currency,
        payer_name=req.payer_name, payment_method=req.payment_method,
        remark=req.remark, entrant_id=current_user.id,
    )
    db.add(r); await db.flush(); return {"id": r.id, "message": "录入成功"}

@router.post("/batch-import")
async def batch_import(req: IncomingBatchImport, current_user: User = Depends(get_current_user),
                       db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "仅仓库管理员可操作")
    imported = 0
    for rec in req.records:
        r = IncomingFlow(
            warehouse_id=rec.get("warehouse_id", get_wh_id(current_user)),
            received_date=datetime.fromisoformat(rec["received_date"]),
            amount=rec["amount"], currency=rec.get("currency", "THB"),
            payer_name=rec.get("payer_name"), payment_method=rec.get("payment_method"),
            remark=rec.get("remark"), entrant_id=current_user.id,
        )
        db.add(r); imported += 1
    await db.flush(); return {"imported": imported, "message": f"成功导入{imported}条记录"}

@router.get("/template")
async def download_template(current_user: User = Depends(get_current_user)):
    """下载到账流水导入模板"""
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    wb = Workbook(); ws = wb.active; ws.title = "到账流水导入模板"
    hf = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    headers = ["到账日期", "金额", "币种", "付款方", "付款方式", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hfont; cell.fill = hf
    # Example row
    example = ["2026-07-01", 5000, "THB", "示例客户", "银行转账", "示例备注"]
    ef = Font(italic=True, color="999999")
    for c, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=c, value=v); cell.font = ef
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    import io; output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=incoming_template.xlsx"})
