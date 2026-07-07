from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.database import get_db
from app.models.supplier import Supplier, SupplierCategory, SupplierProduct, SupplierLogisticsPrice, SupplierCrossBorderPrice
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, Role
from app.schemas.business import SupplierCreate, SupplierUpdate, SupplierResponse, SupplierProductCreate
import io

router = APIRouter()

# ═══ Category CRUD ═══════════════════════════════
@router.get("/categories")
async def list_categories(current_user = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    cats = (await db.execute(select(SupplierCategory).where(SupplierCategory.is_active == "true").order_by(SupplierCategory.sort_order))).scalars().all()
    return {"data": [{"id": c.id, "name": c.name} for c in cats]}

@router.post("/categories")
async def create_category(name: str, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    c = SupplierCategory(name=name)
    db.add(c); await db.flush()
    return {"id": c.id, "name": c.name, "message": "创建成功"}

# ═══ Import Templates ════════════════════════════
@router.get("/import-template/products")
async def download_products_template():
    """下载耗材产品导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "耗材产品导入"
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(["供应商名称", "产品名", "产品规格", "规格报价", "单价"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hfont; cell.fill = hfill
    for c, v in enumerate(["示例: 耗材商A", "快递袋", "一打80个", 80, 1], 1):
        ws.cell(row=2, column=c, value=v)
    for col, w in [('A',20),('B',16),('C',16),('D',12),('E',12)]:
        ws.column_dimensions[col].width = w
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=products_import_template.xlsx"})

@router.get("/import-template/logistics")
async def download_logistics_template():
    """下载跨境物流价格导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "跨境物流价格导入"
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(["供应商名称", "运输方式", "货物类型", "发货仓库", "单价(元/方)", "时效", "币种"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hfont; cell.fill = hfill
    for c, v in enumerate(["示例: 物流公司A", "陆运", "普货", "深圳仓", 800, "5-7天", "人民币"], 1):
        ws.cell(row=2, column=c, value=v)
    for col, w in [('A',22),('B',10),('C',10),('D',10),('E',12),('F',10),('G',8)]:
        ws.column_dimensions[col].width = w
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=cross_border_logistics_template.xlsx"})

# ═══ Import ══════════════════════════════════════
@router.post("/import/products")
async def import_products(file: UploadFile = File(...), current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    """批量导入耗材产品 Excel"""
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    from openpyxl import load_workbook
    content = await file.read()
    wb = load_workbook(io.BytesIO(content)); ws = wb.active
    imported, skipped, errors = 0, 0, []
    # Resolve supplier name→id for this warehouse
    sups = (await db.execute(select(Supplier).where(Supplier.warehouse_id == get_wh_id(current_user)))).scalars().all()
    sup_map = {s.name.strip(): s.id for s in sups}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        sup_name = str(row[0]).strip()
        if sup_name.startswith("示例"): continue
        if sup_name not in sup_map:
            errors.append(f"供应商「{sup_name}」不存在"); skipped += 1; continue
        try:
            p = SupplierProduct(
                supplier_id=sup_map[sup_name],
                product_name=str(row[1] or "").strip(),
                spec=str(row[2] or "").strip() if row[2] else None,
                spec_price=float(row[3] or 0) if row[3] else None,
                unit_price=float(row[4] or 0),
            )
            db.add(p); imported += 1
        except Exception as e:
            errors.append(f"行解析失败: {e}"); skipped += 1
    await db.flush()
    return {"imported": imported, "skipped": skipped, "errors": errors, "message": f"成功导入 {imported} 条产品"}

@router.post("/import/logistics")
async def import_logistics(file: UploadFile = File(...), current_user: User = Depends(get_current_user),
                            db: AsyncSession = Depends(get_db)):
    """批量导入跨境物流价格 Excel"""
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    from openpyxl import load_workbook
    data = await file.read()
    wb = load_workbook(io.BytesIO(data)); ws = wb.active
    imported, skipped, errors = 0, 0, []
    sups = (await db.execute(select(Supplier).where(Supplier.warehouse_id == get_wh_id(current_user)))).scalars().all()
    sup_map = {s.name.strip(): s.id for s in sups}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        sup_name = str(row[0]).strip()
        if sup_name.startswith("示例"): continue
        if sup_name not in sup_map:
            errors.append(f"供应商「{sup_name}」不存在"); skipped += 1; continue
        try:
            p = SupplierCrossBorderPrice(
                supplier_id=sup_map[sup_name],
                transport_method=str(row[1] or "").strip(),
                cargo_type=str(row[2] or "").strip(),
                origin_warehouse=str(row[3] or "").strip(),
                price_per_cbm=float(row[4] or 0),
                estimated_days=str(row[5] or "").strip() if row[5] else None,
                currency=str(row[6] or "人民币").strip(),
            )
            db.add(p); imported += 1
        except Exception as e:
            errors.append(f"行解析失败: {e}"); skipped += 1
    await db.flush()
    return {"imported": imported, "skipped": skipped, "errors": errors, "message": f"成功导入 {imported} 条跨境物流价格"}

# ═══ Product CRUD ════════════════════════════════
@router.get("/{supplier_id}/products")
async def list_products(supplier_id: int, current_user = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    prods = (await db.execute(select(SupplierProduct).where(SupplierProduct.supplier_id == supplier_id).order_by(SupplierProduct.created_at.desc()))).scalars().all()
    return {"data": [{"id": p.id, "product_name": p.product_name, "spec": p.spec, "spec_price": p.spec_price, "unit_price": p.unit_price, "unit": p.unit, "remark": p.remark} for p in prods]}

@router.post("/{supplier_id}/products")
async def add_product(supplier_id: int, req: SupplierProductCreate, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    p = SupplierProduct(supplier_id=supplier_id, **req.model_dump())
    db.add(p); await db.flush()
    return {"id": p.id, "message": "添加成功"}

@router.delete("/{supplier_id}/products/{product_id}")
async def delete_product(supplier_id: int, product_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    p = (await db.execute(select(SupplierProduct).where(SupplierProduct.id == product_id, SupplierProduct.supplier_id == supplier_id))).scalar_one_or_none()
    if not p: raise HTTPException(404, "产品不存在")
    await db.delete(p); await db.flush()
    return {"message": "删除成功"}

# ═══ Logistics Price CRUD ════════════════════════
@router.get("/{supplier_id}/logistics-prices")
async def list_logistics_prices(supplier_id: int, current_user = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    prices = (await db.execute(select(SupplierCrossBorderPrice).where(SupplierCrossBorderPrice.supplier_id == supplier_id).order_by(SupplierCrossBorderPrice.transport_method, SupplierCrossBorderPrice.origin_warehouse))).scalars().all()
    return {"data": [{"id": p.id, "transport_method": p.transport_method, "cargo_type": p.cargo_type,
                      "origin_warehouse": p.origin_warehouse, "price_per_cbm": p.price_per_cbm,
                      "estimated_days": p.estimated_days, "currency": p.currency} for p in prices]}

@router.post("/{supplier_id}/logistics-prices")
async def add_logistics_price(supplier_id: int, req: dict, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    p = SupplierCrossBorderPrice(supplier_id=supplier_id, **req)
    db.add(p); await db.flush()
    return {"id": p.id, "message": "添加成功"}

@router.delete("/{supplier_id}/logistics-prices/{price_id}")
async def delete_logistics_price(supplier_id: int, price_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    p = (await db.execute(select(SupplierCrossBorderPrice).where(SupplierCrossBorderPrice.id == price_id, SupplierCrossBorderPrice.supplier_id == supplier_id))).scalar_one_or_none()
    if not p: raise HTTPException(404, "不存在")
    await db.delete(p); await db.flush()
    return {"message": "删除成功"}

# ═══ Price Comparison ════════════════════════════
@router.get("/compare-prices")
async def compare_prices(product_name: str = None, spec: str = None, category_id: int = None,
                          current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """耗材比价：按产品名称/规格对比同类别供应商报价"""
    if current_user.role not in (Role.WAREHOUSE_ADMIN, Role.STAFF):
        raise HTTPException(403, "无权限")
    q = select(SupplierProduct).join(Supplier, SupplierProduct.supplier_id == Supplier.id).where(Supplier.is_active == "true")
    if product_name:
        q = q.where(SupplierProduct.product_name.ilike(f"%{product_name}%"))
    if spec:
        q = q.where(SupplierProduct.spec.ilike(f"%{spec}%"))
    if category_id:
        q = q.where(Supplier.category_id == category_id)
    elif category_id == 0:
        pass
    else:
        # Default: consumables (category 1)
        pass
    if current_user.role != Role.SUPER_ADMIN:
        q = q.where(Supplier.warehouse_id == get_wh_id(current_user))
    prods = (await db.execute(q.order_by(SupplierProduct.unit_price.asc()))).scalars().all()
    sids = list({p.supplier_id for p in prods})
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap = {s.id: s for s in sups}
    result = []
    for p in prods:
        s = smap.get(p.supplier_id)
        cat_name = ""
        if s and s.category_id:
            cat = (await db.execute(select(SupplierCategory).where(SupplierCategory.id == s.category_id))).scalar_one_or_none()
            if cat: cat_name = cat.name
        result.append({
            "product_id": p.id, "supplier_id": p.supplier_id, "supplier_name": s.name if s else "",
            "category_name": cat_name, "product_name": p.product_name, "spec": p.spec,
            "spec_price": p.spec_price, "unit_price": p.unit_price, "unit": p.unit, "remark": p.remark,
        })
    return {"data": result, "total": len(result)}

@router.get("/compare-logistics")
async def compare_logistics(transport_method: str = None, cargo_type: str = None, origin_warehouse: str = None,
                             category_id: int = None,
                             current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """跨境物流比价：运输方式×货物类型×发货仓库，含义乌加价和最低消费"""
    if current_user.role not in (Role.WAREHOUSE_ADMIN, Role.STAFF):
        raise HTTPException(403, "无权限")
    # 海运最低0.5方，陆运最低0.3方
    MIN_CBM = {"海运": 0.5, "陆运": 0.3}
    YIWU_MARKUP = 120  # 义乌仓加价每方120元
    HEAVY_CARGO_THRESHOLD = 500  # 单方超500kg算重货

    q = select(SupplierCrossBorderPrice).join(Supplier, SupplierCrossBorderPrice.supplier_id == Supplier.id).where(Supplier.is_active == "true")
    if transport_method:
        q = q.where(SupplierCrossBorderPrice.transport_method == transport_method)
    if cargo_type:
        q = q.where(SupplierCrossBorderPrice.cargo_type == cargo_type)
    # 查询时：如果筛选义乌仓，同时查深圳仓和广州仓（做加价计算用）
    is_yiwu = origin_warehouse and "义乌" in origin_warehouse
    if origin_warehouse:
        if is_yiwu:
            # 查义乌+深圳+广州的所有报价
            q = q.where(SupplierCrossBorderPrice.origin_warehouse.in_(["深圳仓", "广州仓", "义乌仓"]))
        else:
            q = q.where(SupplierCrossBorderPrice.origin_warehouse == origin_warehouse)
    if category_id:
        q = q.where(Supplier.category_id == category_id)
    if current_user.role != Role.SUPER_ADMIN:
        q = q.where(Supplier.warehouse_id == get_wh_id(current_user))
    prices = (await db.execute(q.order_by(SupplierCrossBorderPrice.price_per_cbm.asc()))).scalars().all()
    sids = list({p.supplier_id for p in prices})
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap = {s.id: s for s in sups}

    # 构建按 supplier 分组的价格映射
    # key: (supplier_id, transport_method, cargo_type)
    price_map = {}  # {(sid, tm, ct): {warehouse: price}}
    for p in prices:
        key = (p.supplier_id, p.transport_method, p.cargo_type)
        if key not in price_map:
            price_map[key] = {}
        price_map[key][p.origin_warehouse] = {"price_per_cbm": p.price_per_cbm, "price_id": p.id, "estimated_days": p.estimated_days}

    # Build results
    min_cbm = MIN_CBM.get(transport_method or "陆运", 0.3)
    seen = set()
    result = []
    for p in prices:
        key = (p.supplier_id, p.transport_method, p.cargo_type)
        if key in seen: continue
        seen.add(key)
        s = smap.get(p.supplier_id)
        if not s: continue

        # Determine final price
        final_price = p.price_per_cbm
        price_note = ""
        actual_warehouse = p.origin_warehouse

        if is_yiwu:
            pw_map = price_map.get(key, {})
            if "义乌仓" in pw_map:
                final_price = pw_map["义乌仓"]["price_per_cbm"]
                actual_warehouse = "义乌仓"
            elif "深圳仓" in pw_map:
                final_price = pw_map["深圳仓"]["price_per_cbm"] + YIWU_MARKUP
                actual_warehouse = "义乌仓(深圳仓+120)"
                price_note = f"无义乌仓报价，用深圳仓价格+{YIWU_MARKUP}元/方"
            elif "广州仓" in pw_map:
                final_price = pw_map["广州仓"]["price_per_cbm"] + YIWU_MARKUP
                actual_warehouse = "义乌仓(广州仓+120)"
                price_note = f"无义乌仓报价，用广州仓价格+{YIWU_MARKUP}元/方"

        min_amount = round(final_price * min_cbm, 2)
        cat_name = ""
        if s.category_id:
            cat = (await db.execute(select(SupplierCategory).where(SupplierCategory.id == s.category_id))).scalar_one_or_none()
            if cat: cat_name = cat.name
        result.append({
            "price_id": p.id, "supplier_id": p.supplier_id, "supplier_name": s.name,
            "category_name": cat_name,
            "transport_method": p.transport_method, "cargo_type": p.cargo_type,
            "origin_warehouse": actual_warehouse,
            "price_per_cbm": final_price, "price_note": price_note,
            "min_cbm": min_cbm, "min_amount": min_amount,
            "estimated_days": p.estimated_days, "currency": p.currency,
            "heavy_cargo_warning": "单方超500kg按重货计费" if transport_method == "陆运" else "",
        })

    # Sort by final price
    result.sort(key=lambda x: x["price_per_cbm"])
    return {"data": result, "total": len(result), "yiwu_markup": YIWU_MARKUP, "min_cbm": min_cbm}

# ═══ AI Price Analysis ═══════════════════════════
@router.post("/ai-compare")
async def ai_compare(data: dict, current_user: User = Depends(get_current_user)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    from app.config import get_settings
    settings = get_settings()
    if not settings.DEEPSEEK_API_KEY:
        return {"result": "DeepSeek API Key 未配置，无法执行AI比价分析"}
    compare_data = data.get("compare_data", [])
    mode = data.get("mode", "product")  # product or logistics
    if not compare_data:
        return {"result": "无比价数据，请先执行比价查询"}
    if mode == "logistics":
        rows_text = "\n".join([f"- {r['supplier_name']}: {r['route_name']} {r['cargo_type']} 起步价{r['starting_price']} 每公斤{r['price_per_kg']} 时效{r.get('estimated_days','?')}" for r in compare_data])
    else:
        rows_text = "\n".join([f"- {r['supplier_name']}: {r['product_name']} {r['spec'] or ''} 单价{r['unit_price']}{r.get('unit','个')}" for r in compare_data])
    prompt = f"""请对以下供应商报价进行比价分析：
{rows_text}

请从以下维度分析：
1. 价格排名：从低到高列出各供应商
2. 性价比评估：综合考虑价格和已知信息
3. 推荐结论：建议选择哪家供应商及理由"""
    try:
        import httpx
        from openai import AsyncOpenAI
        client = AsyncOpenAI(
            api_key=settings.DEEPSEEK_API_KEY,
            base_url=settings.DEEPSEEK_BASE_URL,
            http_client=httpx.AsyncClient(timeout=60.0),
        )
        resp = await client.chat.completions.create(model="deepseek-chat", messages=[{"role":"user","content": prompt}])
        return {"result": resp.choices[0].message.content}
    except Exception as e:
        return {"result": f"AI分析失败: {str(e)}"}

# ═══ Supplier CRUD ══════════════════════════════
@router.get("")
async def list_suppliers(page: int = 1, page_size: int = 20, search: str = None,
                         category_id: int = None,
                         current_user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(Supplier); count_q = select(func.count(Supplier.id))
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(Supplier.warehouse_id == get_wh_id(current_user))
        count_q = count_q.where(Supplier.warehouse_id == get_wh_id(current_user))
    if search:
        query = query.where(Supplier.name.ilike(f"%{search}%")); count_q = count_q.where(Supplier.name.ilike(f"%{search}%"))
    if category_id:
        query = query.where(Supplier.category_id == category_id)
        count_q = count_q.where(Supplier.category_id == category_id)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(Supplier.created_at.desc()).offset((page-1)*page_size).limit(page_size))
    suppliers = result.scalars().all()
    cat_ids = {s.category_id for s in suppliers if s.category_id}
    cat_map = {}
    if cat_ids:
        cats = (await db.execute(select(SupplierCategory).where(SupplierCategory.id.in_(cat_ids)))).scalars().all()
        cat_map = {c.id: c.name for c in cats}
    return {"data": [{"id": s.id, "name": s.name, "contact_person": s.contact_person, "contact_info": s.contact_info,
                      "address": s.address, "payment_terms": s.payment_terms,
                      "cooperation_content": s.cooperation_content,
                      "settlement_cycle": s.settlement_cycle,
                      "history_notes": s.history_notes,
                      "ai_evaluation": s.ai_evaluation,
                      "category_id": s.category_id,
                      "category_name": cat_map.get(s.category_id, ""),
                      "is_active": s.is_active} for s in suppliers],
            "total": total, "page": page, "page_size": page_size}

@router.post("")
async def create_supplier(req: SupplierCreate, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    d = req.model_dump()
    s = Supplier(warehouse_id=get_wh_id(current_user), **d)
    db.add(s); await db.flush(); return {"id": s.id, "message": "创建成功"}

@router.put("/{supplier_id}")
async def update_supplier(supplier_id: int, req: SupplierUpdate,
                          current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    s = result.scalar_one_or_none()
    if not s: raise HTTPException(404, "供应商不存在")
    if current_user.role != Role.SUPER_ADMIN and s.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "只能修改自己仓库的供应商")
    for k, v in req.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    await db.flush(); return {"message": "更新成功"}

@router.delete("/{supplier_id}")
async def delete_supplier(supplier_id: int, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    s = result.scalar_one_or_none()
    if not s: raise HTTPException(404, "供应商不存在")
    if current_user.role != Role.SUPER_ADMIN and s.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "只能删除自己仓库的供应商")
    await db.delete(s); await db.flush()
    return {"message": "删除成功"}

@router.get("/procurement-summary")
async def procurement_summary(current_user: User = Depends(get_current_user),
                               db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    from app.models.payable import PayableBill
    from app.models.warehouse import Warehouse
    from datetime import date, timedelta
    from sqlalchemy import extract

    wh_id = get_wh_id(current_user) if current_user.role != Role.SUPER_ADMIN else None
    today = date.today()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)

    def wh_filter(q):
        if wh_id: return q.where(PayableBill.warehouse_id == wh_id)
        return q

    # --- 本月支出 ---
    month_q = wh_filter(select(func.coalesce(func.sum(PayableBill.amount), 0)))
    month_total = float((await db.execute(month_q.where(extract("year",PayableBill.bill_date)==today.year, extract("month",PayableBill.bill_date)==today.month))).scalar() or 0)

    # --- 上月支出（同环比） ---
    last_q = wh_filter(select(func.coalesce(func.sum(PayableBill.amount), 0)))
    last_total = float((await db.execute(last_q.where(extract("year",PayableBill.bill_date)==last_month_start.year, extract("month",PayableBill.bill_date)==last_month_start.month))).scalar() or 0)

    # --- 本月耗材/物流分类支出 ---
    cat_month_q = wh_filter(select(Supplier.category_id, func.sum(PayableBill.amount).label("total")).join(Supplier, PayableBill.supplier_id==Supplier.id).where(extract("year",PayableBill.bill_date)==today.year, extract("month",PayableBill.bill_date)==today.month).group_by(Supplier.category_id))
    cat_rows = (await db.execute(cat_month_q)).all()
    cat_map = {}
    cat_ids = [r.category_id for r in cat_rows if r.category_id]
    if cat_ids:
        cats = (await db.execute(select(SupplierCategory).where(SupplierCategory.id.in_(cat_ids)))).scalars().all()
        cat_map = {c.id: c.name for c in cats}
    cat_spending = {}
    for r in cat_rows:
        name = cat_map.get(r.category_id, "未分类")
        cat_spending[name] = float(r.total or 0)

    # --- 环比 ---
    pct_change = 0
    if last_total > 0:
        pct_change = round((month_total - last_total) / last_total * 100, 1)

    # --- 供应商排名 ---
    sup_q = wh_filter(select(PayableBill.supplier_id, func.sum(PayableBill.amount).label("total"), func.max(PayableBill.bill_date).label("last_date")).group_by(PayableBill.supplier_id).order_by(func.sum(PayableBill.amount).desc()))
    sup_rows = (await db.execute(sup_q)).all()
    # 当月支出
    sup_month_q = wh_filter(select(PayableBill.supplier_id, func.sum(PayableBill.amount).label("total")).where(extract("year",PayableBill.bill_date)==today.year, extract("month",PayableBill.bill_date)==today.month).group_by(PayableBill.supplier_id))
    sup_month = {r.supplier_id: float(r.total or 0) for r in (await db.execute(sup_month_q)).all()}
    sids = [r.supplier_id for r in sup_rows]
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        cat_ids2 = {s.category_id for s in sups if s.category_id}
        cat2 = {}
        if cat_ids2:
            crows = (await db.execute(select(SupplierCategory).where(SupplierCategory.id.in_(cat_ids2)))).scalars().all()
            cat2 = {c.id: c.name for c in crows}
        smap = {s.id: {"name": s.name, "category": cat2.get(s.category_id, "")} for s in sups}
    supplier_ranking = []
    for r in sup_rows:
        info = smap.get(r.supplier_id, {"name": "", "category": ""})
        supplier_ranking.append({
            "supplier_id": r.supplier_id, "supplier_name": info["name"],
            "category_name": info["category"],
            "month_amount": sup_month.get(r.supplier_id, 0),
            "total_amount": float(r.total or 0),
            "last_bill_date": r.last_date.isoformat()[:10] if r.last_date else None,
        })

    # --- 产品比价汇总 ---
    prod_q = wh_filter(select(SupplierProduct.product_name, SupplierProduct.spec,
        func.count(SupplierProduct.supplier_id.distinct()).label("supplier_count"),
        func.min(SupplierProduct.unit_price).label("min_price"),
        func.max(SupplierProduct.unit_price).label("max_price"))
        .join(Supplier, SupplierProduct.supplier_id==Supplier.id)
        .group_by(SupplierProduct.product_name, SupplierProduct.spec)
        .order_by(SupplierProduct.product_name))
    prod_rows = (await db.execute(prod_q)).all()
    product_compare = []
    for p in prod_rows:
        if not p.product_name: continue
        # Find which supplier has the min price
        min_sup_q = wh_filter(select(SupplierProduct.supplier_id, Supplier.name)
            .join(Supplier, SupplierProduct.supplier_id==Supplier.id)
            .where(SupplierProduct.product_name==p.product_name, SupplierProduct.spec==p.spec, SupplierProduct.unit_price==p.min_price))
        min_sup = (await db.execute(min_sup_q)).first()
        product_compare.append({
            "product_name": p.product_name, "spec": p.spec,
            "supplier_count": p.supplier_count,
            "min_price": float(p.min_price) if p.min_price else 0,
            "max_price": float(p.max_price) if p.max_price else 0,
            "min_supplier": min_sup[1] if min_sup else "",
        })

    # --- 省钱提示：对比同类产品不同供应商 ---
    savings_tips = []
    for prod in product_compare:
        if prod["supplier_count"] >= 2 and prod["max_price"] > prod["min_price"]:
            diff = round(prod["max_price"] - prod["min_price"], 2)
            savings_tips.append({
                "product_name": prod["product_name"], "spec": prod["spec"],
                "cheapest_price": prod["min_price"], "cheapest_supplier": prod["min_supplier"],
                "highest_price": prod["max_price"],
                "savings_per_unit": diff,
                "tip": f"{prod['product_name']}{prod['spec'] or ''}：最便宜 {prod['min_supplier']} ¥{prod['min_price']}，最贵 ¥{prod['max_price']}，用便宜的可省 ¥{diff}/件",
            })

    return {
        "overview": {
            "month_total": month_total, "last_month_total": last_total,
            "pct_change": pct_change,
            "cat_spending": cat_spending,
        },
        "supplier_ranking": supplier_ranking,
        "product_compare": product_compare,
        "savings_tips": savings_tips,
    }

@router.get("/{supplier_id}")
async def get_supplier(supplier_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    s = result.scalar_one_or_none()
    if not s: raise HTTPException(404, "供应商不存在")
    if current_user.role != Role.SUPER_ADMIN and s.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "只能查看自己仓库的供应商")
    cat_name = ""
    if s.category_id:
        cat = (await db.execute(select(SupplierCategory).where(SupplierCategory.id == s.category_id))).scalar_one_or_none()
        if cat: cat_name = cat.name
    return {"id": s.id, "name": s.name, "contact_person": s.contact_person, "contact_info": s.contact_info,
            "address": s.address, "payment_terms": s.payment_terms,
            "cooperation_content": s.cooperation_content,
            "settlement_cycle": s.settlement_cycle,
            "history_notes": s.history_notes,
            "ai_evaluation": s.ai_evaluation,
            "category_id": s.category_id, "category_name": cat_name}

@router.get("/{supplier_id}/ai-evaluation")
async def ai_evaluate(supplier_id: int, current_user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "仅超级管理员可用")
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    s = result.scalar_one_or_none()
    if not s: raise HTTPException(404, "供应商不存在")
    from app.config import get_settings
    settings = get_settings()
    if settings.DEEPSEEK_API_KEY:
        try:
            import httpx
            from openai import AsyncOpenAI
            client = AsyncOpenAI(
                api_key=settings.DEEPSEEK_API_KEY,
                base_url=settings.DEEPSEEK_BASE_URL,
                http_client=httpx.AsyncClient(timeout=60.0),
            )
            resp = await client.chat.completions.create(model="deepseek-chat", messages=[{"role":"user",
                "content":f"评估供应商：{s.name}，联系方式：{s.contact_info or '无'}，地址：{s.address or '无'}。请从信誉、价格竞争力、交付及时性三个维度简要评估。"}])
            evaluation = {"result": resp.choices[0].message.content}
            s.ai_evaluation = evaluation
        except Exception as e:
            evaluation = {"error": str(e)}
    else:
        evaluation = {"result": "DeepSeek API Key 未配置，无法执行AI评估"}
    if not s.ai_evaluation:
        s.ai_evaluation = evaluation
    await db.flush()
    return evaluation
