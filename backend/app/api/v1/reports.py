from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime
from app.database import get_db
from app.models.recharge import RechargeDeclaration, IncomingFlow, ReconciliationResult
from app.models.income_expense import IncomeRecord, ExpenseRecord
from app.models.expense_fund import ExpenseFund
from app.models.reimbursement import Reimbursement
from app.models.payable import PayableBill
from app.models.credit import CreditCustomer
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, get_wh_ids, Role
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()

STATUS_LABELS = {
    "pending": "待处理", "matched": "已匹配", "unmatched": "未匹配",
    "paid": "已付款", "active": "正常", "approved": "已通过",
    "rejected": "已驳回", "partially_approved": "部分通过",
    "fund_linked": "转入备用金审核", "partially_paid": "部分付款",
    "overdue": "逾期", "active": "正常", "paused": "暂停", "cancelled": "已取消",
    "settled": "已结清",
}

def _(s): return STATUS_LABELS.get(s, s)

# Map Chinese sheet names to safe English filenames
_FILENAME_MAP = {
    "充值汇总": "recharge_summary", "到账汇总": "incoming_summary",
    "收支报表": "income_expense", "应付报表": "payable",
    "备用金报表": "expense_fund", "报销报表": "reimbursement",
    "账期报表": "credit", "对账差异": "reconciliation_diff",
}
def to_excel(headers, rows, sheet_name="Sheet1"):
    wb = Workbook(); ws = wb.active; ws.title = sheet_name
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fw = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = fw; c.fill = fill; c.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1): ws.cell(row=r, column=c, value=val)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    safe_name = _FILENAME_MAP.get(sheet_name, sheet_name.replace(" ", "_"))
    filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    # Use RFC 5987 encoding for safe filename delivery
    from urllib.parse import quote
    encoded = quote(filename)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"})



# ===== 0. 报表预览（所有卡片一次性返回） =====
@router.get("/previews")
async def report_previews(
    month: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    wh_ids = get_wh_ids(current_user)
    today = __import__('datetime').date.today()
    m = month or f"{today.year}-{today.month:02d}"

    previews = {}

    # 1. 充值汇总
    q = select(func.coalesce(func.sum(RechargeDeclaration.amount), 0), func.count(RechargeDeclaration.id))
    if wh_ids: q = q.where(RechargeDeclaration.warehouse_id.in_(wh_ids))
    if m: q = q.where(func.to_char(RechargeDeclaration.declare_date, 'YYYY-MM') == m)
    total_amt, count = (await db.execute(q)).first()
    previews["recharge-summary"] = {"preview": float(total_amt or 0), "count": count or 0}

    # 2. 到账汇总
    q = select(func.coalesce(func.sum(IncomingFlow.amount), 0), func.count(IncomingFlow.id))
    if wh_ids: q = q.where(IncomingFlow.warehouse_id.in_(wh_ids))
    if m: q = q.where(func.to_char(IncomingFlow.received_date, 'YYYY-MM') == m)
    total_amt, count = (await db.execute(q)).first()
    previews["incoming-summary"] = {"preview": float(total_amt or 0), "count": count or 0}

    # 3. 收支报表
    iq = select(func.coalesce(func.sum(IncomeRecord.amount), 0))
    eq = select(func.coalesce(func.sum(ExpenseRecord.amount), 0))
    if wh_ids: iq = iq.where(IncomeRecord.warehouse_id.in_(wh_ids)); eq = eq.where(ExpenseRecord.warehouse_id.in_(wh_ids))
    if m: iq = iq.where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == m); eq = eq.where(func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == m)
    inc = float((await db.execute(iq)).scalar() or 0)
    exp = float((await db.execute(eq)).scalar() or 0)
    # Also add recharge income
    rq2 = select(func.coalesce(func.sum(RechargeDeclaration.amount), 0))
    if wh_ids: rq2 = rq2.where(RechargeDeclaration.warehouse_id.in_(wh_ids))
    if m: rq2 = rq2.where(func.to_char(RechargeDeclaration.declare_date, 'YYYY-MM') == m)
    rch = float((await db.execute(rq2)).scalar() or 0)
    previews["income-expense"] = {"preview": inc + exp + rch, "total_income": inc + rch, "total_expense": exp, "net": inc + rch - exp}

    # 4. 应付报表
    q = select(func.coalesce(func.sum(PayableBill.amount - func.coalesce(PayableBill.paid_amount, 0)), 0), func.count(PayableBill.id))
    if wh_ids: q = q.where(PayableBill.warehouse_id.in_(wh_ids))
    if m: q = q.where(func.to_char(PayableBill.due_date, 'YYYY-MM') == m)
    pending, count = (await db.execute(q)).first()
    previews["payable"] = {"preview": float(pending or 0), "count": count or 0}

    # 5. 备用金报表
    q = select(func.coalesce(func.sum(ExpenseFund.remaining_balance), 0), func.count(ExpenseFund.id))
    if wh_ids: q = q.where(ExpenseFund.warehouse_id.in_(wh_ids))
    q = q.where(ExpenseFund.status == "active")
    in_transit, count = (await db.execute(q)).first()
    previews["expense-fund"] = {"preview": float(in_transit or 0), "count": count or 0}

    # 6. 报销报表
    q = select(func.coalesce(func.sum(Reimbursement.total_amount), 0), func.count(Reimbursement.id))
    if wh_ids: q = q.where(Reimbursement.warehouse_id.in_(wh_ids))
    if m: q = q.where(func.to_char(Reimbursement.submit_date, 'YYYY-MM') == m)
    total_amt, count = (await db.execute(q)).first()
    previews["reimbursement"] = {"preview": float(total_amt or 0), "count": count or 0}

    # 7. 账期报表
    q = select(func.coalesce(func.sum(CreditCustomer.current_debt), 0), func.count(CreditCustomer.id))
    if wh_ids: q = q.where(CreditCustomer.warehouse_id.in_(wh_ids))
    total_debt, count = (await db.execute(q)).first()
    previews["credit"] = {"preview": float(total_debt or 0), "count": count or 0}

    # 8. 对账差异
    q = select(func.coalesce(func.sum(func.abs(ReconciliationResult.amount_diff)), 0), func.count(ReconciliationResult.id))
    if wh_ids: q = q.where(ReconciliationResult.warehouse_id.in_(wh_ids))
    if m: q = q.where(ReconciliationResult.reconciliation_month == m)
    total_diff, count = (await db.execute(q)).first()
    previews["reconciliation-diff"] = {"preview": float(total_diff or 0), "count": count or 0}

    return {"previews": previews, "month": m}

# ===== 1. 充值汇总 =====
@router.get("/recharge-summary")
async def recharge_summary(month: str = None, warehouse_id: int = None, format: str = "json",
                            current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(RechargeDeclaration)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(RechargeDeclaration.warehouse_id.in_(wh_ids))
    if month: query = query.where(func.to_char(RechargeDeclaration.declare_date, 'YYYY-MM') == month)
    if warehouse_id: query = query.where(RechargeDeclaration.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(RechargeDeclaration.declare_date.desc()))
    records = result.scalars().all()
    data = [{"日期": r.declare_date.strftime("%Y-%m-%d") if r.declare_date else "", "金额": r.amount, "币种": r.currency, "状态": _(r.match_status)} for r in records]
    headers = ["日期","金额","币种","状态"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "充值汇总")
    return {"data": data, "total_amount": sum(r.amount for r in records), "total_count": len(records)}

# ===== 2. 到账汇总 =====
@router.get("/incoming-summary")
async def incoming_summary(month: str = None, warehouse_id: int = None, format: str = "json",
                            current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(IncomingFlow)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(IncomingFlow.warehouse_id.in_(wh_ids))
    if month: query = query.where(func.to_char(IncomingFlow.received_date, 'YYYY-MM') == month)
    if warehouse_id: query = query.where(IncomingFlow.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(IncomingFlow.received_date.desc()))
    records = result.scalars().all()
    data = [{"日期": r.received_date.strftime("%Y-%m-%d") if r.received_date else "", "金额": r.amount, "币种": r.currency, "付款方": r.payer_name or ""} for r in records]
    headers = ["日期","金额","币种","付款方"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "到账汇总")
    return {"data": data, "total_amount": sum(r.amount for r in records), "total_count": len(records)}

# ===== 3. 收支报表 =====
@router.get("/income-expense")
async def income_expense_report(month: str = None, warehouse_id: int = None, format: str = "json",
                                 current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    wh_ids = get_wh_ids(current_user)
    iq = select(IncomeRecord); eq = select(ExpenseRecord)
    rq = select(RechargeDeclaration)  # 充值申报也作为收入来源
    if wh_ids: iq = iq.where(IncomeRecord.warehouse_id.in_(wh_ids)); eq = eq.where(ExpenseRecord.warehouse_id.in_(wh_ids))
    if wh_ids: rq = rq.where(RechargeDeclaration.warehouse_id.in_(wh_ids))
    if month: iq = iq.where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == month); eq = eq.where(func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == month)
    if month: rq = rq.where(func.to_char(RechargeDeclaration.declare_date, 'YYYY-MM') == month)
    if warehouse_id: iq = iq.where(IncomeRecord.warehouse_id == warehouse_id); eq = eq.where(ExpenseRecord.warehouse_id == warehouse_id)
    if warehouse_id: rq = rq.where(RechargeDeclaration.warehouse_id == warehouse_id)

    incomes = (await db.execute(iq.order_by(IncomeRecord.income_date.desc()))).scalars().all()
    expenses = (await db.execute(eq.order_by(ExpenseRecord.expense_date.desc()))).scalars().all()
    recharges = (await db.execute(rq.order_by(RechargeDeclaration.declare_date.desc()))).scalars().all()

    # 获取充值对应的客户名称
    recharge_cust_ids = list(set(r.customer_id for r in recharges if r.customer_id))
    recharge_cust_map = {}
    if recharge_cust_ids:
        from app.models.customer import Customer
        custs = (await db.execute(select(Customer).where(Customer.id.in_(recharge_cust_ids)))).scalars().all()
        recharge_cust_map = {c.id: c.company_name for c in custs}

    data = []
    for r in incomes:
        data.append({"类型": "收入", "日期": r.income_date.strftime("%Y-%m-%d") if r.income_date else "", "金额": r.amount, "币种": r.currency or "THB", "备注": r.remark or ""})
    for r in recharges:
        cust_name = recharge_cust_map.get(r.customer_id, "") if r.customer_id else ""
        label = f"充值-{cust_name}" if cust_name else "充值收入"
        data.append({"类型": "充值收入", "日期": r.declare_date.strftime("%Y-%m-%d") if r.declare_date else "", "金额": r.amount, "币种": r.currency or "THB", "备注": label})
    for r in expenses:
        data.append({"类型": "支出", "日期": r.expense_date.strftime("%Y-%m-%d") if r.expense_date else "", "金额": r.amount, "币种": r.currency or "THB", "备注": r.remark or ""})
    data.sort(key=lambda x: x["日期"], reverse=True)

    total_income = sum(r.amount for r in incomes) + sum(r.amount for r in recharges)
    total_expense = sum(r.amount for r in expenses)
    recharge_income = sum(r.amount for r in recharges)
    other_income = sum(r.amount for r in incomes)
    headers = ["类型","日期","金额","币种","备注"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "收支报表")
    return {"data": data, "total_income": total_income, "total_expense": total_expense, "net": total_income - total_expense,
            "recharge_income": recharge_income, "other_income": other_income}

# ===== 4. 应付报表 =====
@router.get("/payable")
async def payable_report(month: str = None, warehouse_id: int = None, format: str = "json",
                          current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PayableBill)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(PayableBill.warehouse_id.in_(wh_ids))
    if month: query = query.where(func.to_char(PayableBill.due_date, 'YYYY-MM') == month)
    if warehouse_id: query = query.where(PayableBill.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(PayableBill.due_date))
    bills = result.scalars().all()
    data = [{"账单号": b.bill_number or "", "到期日": b.due_date.strftime("%Y-%m-%d") if b.due_date else "", "金额": b.amount or 0,
             "已付": b.paid_amount or 0, "状态": _(b.status)} for b in bills]
    pending_total = sum(b.amount - (b.paid_amount or 0) for b in bills if b.status in ("pending","partially_paid","overdue"))
    overdue_count = sum(1 for b in bills if b.status == "overdue")
    headers = ["账单号","到期日","金额","已付","状态"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "应付报表")
    return {"data": data, "total_pending": pending_total, "overdue_count": overdue_count, "total_count": len(bills)}

# ===== 5. 备用金报表 =====
@router.get("/expense-fund")
async def expense_fund_report(warehouse_id: int = None, format: str = "json",
                               current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(ExpenseFund)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(ExpenseFund.warehouse_id.in_(wh_ids))
    if warehouse_id: query = query.where(ExpenseFund.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(ExpenseFund.created_at.desc()))
    funds = result.scalars().all()
    data = [{"用途": f.purpose or "", "金额": f.amount or 0, "余额": f.remaining_balance or 0, "状态": _(f.status)} for f in funds]
    in_transit = sum(f.remaining_balance or 0 for f in funds if f.status == "active")
    headers = ["用途","金额","余额","状态"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "备用金报表")
    return {"data": data, "in_transit_total": in_transit, "total_count": len(funds)}

# ===== 6. 报销报表 =====
@router.get("/reimbursement")
async def reimbursement_report(month: str = None, warehouse_id: int = None, format: str = "json",
                                current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(Reimbursement)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(Reimbursement.warehouse_id.in_(wh_ids))
    if month: query = query.where(func.to_char(Reimbursement.submit_date, 'YYYY-MM') == month)
    if warehouse_id: query = query.where(Reimbursement.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(Reimbursement.submit_date.desc()))
    reimbs = result.scalars().all()
    data = [{"日期": r.submit_date.strftime("%Y-%m-%d") if r.submit_date else "", "金额": r.total_amount or 0,
             "币种": r.currency or "THB", "状态": _(r.status)} for r in reimbs]
    total = sum(r.total_amount or 0 for r in reimbs)
    headers = ["日期","金额","币种","状态"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "报销报表")
    return {"data": data, "total_amount": total, "total_count": len(reimbs)}

# ===== 7. 账期报表 =====
@router.get("/credit")
async def credit_report(warehouse_id: int = None, format: str = "json",
                         current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(CreditCustomer)
    wh_ids = get_wh_ids(current_user)
    if wh_ids: query = query.where(CreditCustomer.warehouse_id.in_(wh_ids))
    if warehouse_id: query = query.where(CreditCustomer.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(CreditCustomer.created_at.desc()))
    credits = result.scalars().all()
    data = [{"额度": c.credit_limit or 0, "欠款": c.current_debt or 0,
             "逾期天数": c.overdue_days or 0, "状态": _(c.status)} for c in credits]
    total_debt = sum(c.current_debt or 0 for c in credits)
    overdue_count = sum(1 for c in credits if (c.overdue_days or 0) > 0)
    headers = ["额度","欠款","逾期天数","状态"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "账期报表")
    return {"data": data, "total_debt": total_debt, "overdue_count": overdue_count}

# ===== 8. 对账差异 =====
@router.get("/reconciliation-diff")
async def reconciliation_diff_report(month: str, warehouse_id: int = None, format: str = "json",
                                      current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    wh_ids = get_wh_ids(current_user)
    query = select(ReconciliationResult).where(ReconciliationResult.reconciliation_month == month)
    if wh_ids: query = query.where(ReconciliationResult.warehouse_id.in_(wh_ids))
    if warehouse_id: query = query.where(ReconciliationResult.warehouse_id == warehouse_id)
    result = await db.execute(query.order_by(ReconciliationResult.id))
    records = result.scalars().all()
    data = [{"月份": r.reconciliation_month or "", "状态": _(r.match_status), "差额": r.amount_diff or 0, "处理说明": r.handling_note or ""} for r in records]
    diff_count = sum(1 for r in records if (r.amount_diff or 0) != 0)
    total_diff = sum(abs(r.amount_diff or 0) for r in records)
    headers = ["月份","状态","差额","处理说明"]
    if format == "excel": return to_excel(headers, [[r[h] for h in headers] for r in data], "对账差异")
    return {"data": data, "diff_count": diff_count, "total_diff": total_diff}
