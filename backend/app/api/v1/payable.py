from fastapi.responses import StreamingResponse
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime
from app.database import get_db
from app.models.payable import PayableBill, PayablePlan, PlanTemplate, PayableStatus, PlanStatus, MonthlyOrderVolume
from app.models.supplier import Supplier, SupplierCategory
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, Role
from pydantic import BaseModel
from typing import Optional, List

router = APIRouter()

class BillCreate(BaseModel):
    supplier_id: int; bill_number: str; bill_date: str; due_date: str
    amount: float; confirmed_amount: Optional[float] = None
    currency: str = "THB"; remark: Optional[str] = None
    payment_commitment_days: Optional[int] = None
    detail: Optional[str] = None
    is_fund_linked: Optional[str] = None

class PlanCreate(BaseModel):
    plan_name: str; planned_date: str; bill_ids: List[int]; remark: Optional[str] = None
    save_as_template: bool = False; template_name: Optional[str] = None

@router.get("")
async def list_bills(
    page: int = 1, page_size: int = 20, supplier_id: int = None,
    status: str = None, month: str = None,
    current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PayableBill); count_q = select(func.count(PayableBill.id))
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayableBill.warehouse_id == get_wh_id(current_user))
        count_q = count_q.where(PayableBill.warehouse_id == get_wh_id(current_user))
    if supplier_id:
        query = query.where(PayableBill.supplier_id == supplier_id)
        count_q = count_q.where(PayableBill.supplier_id == supplier_id)
    if status:
        query = query.where(PayableBill.status == status)
        count_q = count_q.where(PayableBill.status == status)
    if month:
        query = query.where(func.to_char(PayableBill.bill_date, 'YYYY-MM') == month)
        count_q = count_q.where(func.to_char(PayableBill.bill_date, 'YYYY-MM') == month)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(PayableBill.due_date.asc()).offset((page-1)*page_size).limit(page_size))
    bills = result.scalars().all()
    sids = {b.supplier_id for b in bills}
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap = {s.id: s.name for s in sups}
    # 自动标记逾期（due_date < now 且未付）
    from sqlalchemy import update as sql_update
    import sqlalchemy
    expire_sql = (
        sql_update(PayableBill)
        .where(PayableBill.due_date < datetime.now())
        .where(PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value]))
        .values(status=PayableStatus.OVERDUE.value)
    )
    try:
        await db.execute(expire_sql)
        await db.flush()
    except: pass

    return {
        "data": [{
            "id": b.id, "warehouse_id": b.warehouse_id, "supplier_id": b.supplier_id,
            "supplier_name": smap.get(b.supplier_id, ""),
            "bill_number": b.bill_number,
            "bill_date": b.bill_date.isoformat() if b.bill_date else None,
            "due_date": b.due_date.isoformat() if b.due_date else None,
            "amount": b.amount, "confirmed_amount": b.confirmed_amount,
            "currency": b.currency, "paid_amount": b.paid_amount,
            "status": b.status, "is_duplicate_warned": b.is_duplicate_warned,
            "payment_commitment_days": b.payment_commitment_days,
            "payment_voucher": b.payment_voucher, "payment_method": b.payment_method, "bill_attachment": b.bill_attachment,
            "is_fund_linked": b.is_fund_linked, "detail": b.detail, "remark": b.remark,
            "diff_note": b.diff_note,
        } for b in bills],
        "total": total, "page": page, "page_size": page_size,
    }

@router.post("")
async def create_bill(req: BillCreate, current_user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    # 重复检测
    existing = (await db.execute(
        select(PayableBill).where(
            PayableBill.warehouse_id == get_wh_id(current_user),
            PayableBill.bill_number == req.bill_number,
            PayableBill.supplier_id == req.supplier_id,
        )
    )).scalar_one_or_none()
    if existing and not existing.is_duplicate_warned:
        raise HTTPException(409, f"重复账单警告: 账单号 {req.bill_number} 已存在，请确认是否新账单")

    # 对账差异检测
    has_diff = False
    diff_note = None
    if req.confirmed_amount is not None and req.confirmed_amount != req.amount:
        has_diff = True
        diff_note = f"供应商确认金额 {req.confirmed_amount} 与仓库记录 {req.amount} 不一致"
    
    b = PayableBill(
        warehouse_id=get_wh_id(current_user), supplier_id=req.supplier_id,
        bill_number=req.bill_number,
        bill_date=datetime.fromisoformat(req.bill_date),
        due_date=datetime.fromisoformat(req.due_date),
        amount=req.amount, confirmed_amount=req.confirmed_amount,
        currency=req.currency, detail=req.detail, remark=req.remark,
        payment_commitment_days=req.payment_commitment_days,
        is_fund_linked=req.is_fund_linked,
        created_by=current_user.id,
    )
    db.add(b); await db.flush()
    return {"id": b.id, "message": "账单创建成功", "has_diff": has_diff, "diff_note": diff_note}

@router.post("/{bill_id}/upload-voucher")
async def upload_voucher(bill_id: int, file: UploadFile = File(...),
                         current_user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    result = await db.execute(select(PayableBill).where(PayableBill.id == bill_id))
    b = result.scalar_one_or_none()
    if not b: raise HTTPException(404, "账单不存在")
    import os, uuid
    upload_dir = "/app/uploads/payment_vouchers"
    os.makedirs(upload_dir, exist_ok=True)
    ext = file.filename.split(".")[-1] if file.filename and "." in file.filename else "png"
    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = os.path.join(upload_dir, fname)
    content = await file.read()
    with open(fpath, "wb") as f: f.write(content)
    b.payment_voucher = f"/uploads/payment_vouchers/{fname}"
    await db.flush()
    return {"message": "凭证上传成功", "path": b.payment_voucher}

@router.post("/{bill_id}/upload-attachment")
async def upload_bill_attachment(bill_id: int, file: UploadFile = File(...),
                                 current_user: User = Depends(get_current_user),
                                 db: AsyncSession = Depends(get_db)):
    if current_user.role not in (Role.WAREHOUSE_ADMIN,):
        raise HTTPException(403, "无权限")
    result = await db.execute(select(PayableBill).where(PayableBill.id == bill_id))
    b = result.scalar_one_or_none()
    if not b: raise HTTPException(404, "账单不存在")
    import os, uuid
    upload_dir = "/app/uploads/bill_attachments"
    os.makedirs(upload_dir, exist_ok=True)
    ext = file.filename.split(".")[-1] if file.filename and "." in file.filename else "pdf"
    fname = f"{uuid.uuid4().hex}.{ext}"
    fpath = os.path.join(upload_dir, fname)
    content = await file.read()
    with open(fpath, "wb") as f: f.write(content)
    b.bill_attachment = f"/uploads/bill_attachments/{fname}"
    await db.flush()
    return {"message": "账单附件上传成功", "path": b.bill_attachment}

@router.put("/{bill_id}/pay")
async def pay_bill(bill_id: int, paid_amount: float = None, payment_method: str = None,
                   current_user: User = Depends(get_current_user),
                   db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    result = await db.execute(select(PayableBill).where(PayableBill.id == bill_id))
    b = result.scalar_one_or_none()
    if not b: raise HTTPException(404, "账单不存在")
    remaining = b.amount - b.paid_amount
    pay = paid_amount if paid_amount is not None else remaining
    # Boundary checks
    if pay <= 0:
        raise HTTPException(400, "付款金额必须大于0")
    if pay > remaining:
        raise HTTPException(400, f"付款金额({pay:,.2f})超过剩余应付({remaining:,.2f})")
    b.paid_amount += pay
    b.paid_at = datetime.now()
    if payment_method:
        b.payment_method = payment_method
    if b.paid_amount >= b.amount:
        b.status = PayableStatus.PAID.value
    else:
        b.status = PayableStatus.PARTIALLY_PAID.value
    # 清除逾期标记
    if b.status in (PayableStatus.PAID.value, PayableStatus.PARTIALLY_PAID.value) and b.status != PayableStatus.OVERDUE.value:
        pass
    await db.flush(); return {"message": "付款记录成功"}

class BillUpdate(BaseModel):
    confirmed_amount: Optional[float] = None
    detail: Optional[str] = None
    remark: Optional[str] = None
    diff_note: Optional[str] = None

@router.put("/{bill_id}")
async def update_bill(bill_id: int, req: BillUpdate,
                      current_user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    result = await db.execute(select(PayableBill).where(PayableBill.id == bill_id))
    b = result.scalar_one_or_none()
    if not b: raise HTTPException(404, "账单不存在")
    if current_user.role != Role.SUPER_ADMIN and b.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "只能编辑自己仓库的账单")
    if req.confirmed_amount is not None:
        b.confirmed_amount = req.confirmed_amount
    if req.detail is not None:
        b.detail = req.detail
    if req.remark is not None:
        b.remark = req.remark
    if req.diff_note is not None:
        b.diff_note = req.diff_note
    await db.flush()
    return {"message": "账单更新成功", "has_diff": b.confirmed_amount is not None and b.confirmed_amount != b.amount}

# === Stats Dashboard ===
@router.get("/stats")
async def payable_stats(current_user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from datetime import date
    from sqlalchemy import extract
    wh_id = get_wh_id(current_user)
    today = date.today()

    def wh(q):
        return q.where(PayableBill.warehouse_id == wh_id)

    # Month totals
    month_cond = [extract("year", PayableBill.bill_date) == today.year, extract("month", PayableBill.bill_date) == today.month]
    month_total = float((await db.execute(wh(select(func.coalesce(func.sum(PayableBill.amount), 0))).where(*month_cond))).scalar() or 0)
    month_paid = float((await db.execute(wh(select(func.coalesce(func.sum(PayableBill.paid_amount), 0))).where(*month_cond))).scalar() or 0)
    month_unpaid = month_total - month_paid

    # Overdue
    overdue_total = float((await db.execute(wh(select(func.coalesce(func.sum(PayableBill.amount - PayableBill.paid_amount), 0)))
        .where(PayableBill.status == PayableStatus.OVERDUE.value))).scalar() or 0)

    # Supplier summary (month)
    sup_q = wh(select(PayableBill.supplier_id,
        func.count(PayableBill.id).label("count"),
        func.sum(PayableBill.amount).label("total"),
        func.sum(PayableBill.paid_amount).label("paid"))
        .where(*month_cond).group_by(PayableBill.supplier_id).order_by(func.sum(PayableBill.amount).desc()))
    sup_rows = (await db.execute(sup_q)).all()
    sids = [r.supplier_id for r in sup_rows]
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap = {s.id: s.name for s in sups}
    supplier_summary = [{
        "supplier_id": r.supplier_id, "supplier_name": smap.get(r.supplier_id, ""),
        "bill_count": r.count, "total_amount": float(r.total or 0),
        "paid_amount": float(r.paid or 0), "unpaid_amount": float((r.total or 0) - (r.paid or 0)),
    } for r in sup_rows]

    return {
        "month_total": month_total, "overdue_total": overdue_total,
        "month_paid": month_paid, "month_unpaid": month_unpaid,
        "supplier_summary": supplier_summary,
    }

# === Plans ===
@router.get("/plans")
async def list_plans(current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PayablePlan)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayablePlan.warehouse_id == get_wh_id(current_user))
    result = await db.execute(query.order_by(PayablePlan.planned_date.desc()))
    plans = result.scalars().all()
    # 自动标记逾期（due_date < now 且未付）
    from sqlalchemy import update as sql_update
    import sqlalchemy
    expire_sql = (
        sql_update(PayableBill)
        .where(PayableBill.due_date < datetime.now())
        .where(PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value]))
        .values(status=PayableStatus.OVERDUE.value)
    )
    try:
        await db.execute(expire_sql)
        await db.flush()
    except: pass

    return {"data": [{"id": p.id, "plan_name": p.plan_name, "planned_date": p.planned_date.isoformat() if p.planned_date else None, "total_amount": p.total_amount, "status": p.status, "bill_ids": p.bill_ids, "remark": p.remark} for p in plans]}

@router.post("/plans")
async def create_plan(req: PlanCreate, current_user: User = Depends(get_current_user),
                      db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    total = 0
    if req.bill_ids:
        bills = (await db.execute(select(PayableBill).where(PayableBill.id.in_(req.bill_ids)))).scalars().all()
        total = sum(b.amount - b.paid_amount for b in bills)
    p = PayablePlan(
        warehouse_id=get_wh_id(current_user), plan_name=req.plan_name,
        planned_date=datetime.fromisoformat(req.planned_date),
        total_amount=total, bill_ids=req.bill_ids, remark=req.remark,
        created_by=current_user.id,
    )
    db.add(p); await db.flush()
    # Save as template if requested
    if req.save_as_template and req.bill_ids:
        tmpl = PlanTemplate(
            warehouse_id=get_wh_id(current_user),
            name=req.template_name or req.plan_name,
            bill_ids=req.bill_ids,
            created_by=current_user.id,
        )
        db.add(tmpl)
    return {"id": p.id, "message": "付款计划创建成功"}

# === Cashflow ===
@router.get("/cashflow-prediction")
async def cashflow_prediction(current_user: User = Depends(get_current_user),
                              db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(func.coalesce(func.sum(PayableBill.amount - PayableBill.paid_amount), 0))
    query = query.where(PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value]))
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayableBill.warehouse_id == get_wh_id(current_user))
    total_pending = (await db.execute(query)).scalar() or 0
    return {"total_pending_payable": float(total_pending), "predicted_outflow": float(total_pending)}

# === Batch export ===
@router.get("/batch-export")
async def batch_export(supplier_id: int = None, start_date: str = None, end_date: str = None,
                       status: str = None, bill_ids: str = None,
                       current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PayableBill)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayableBill.warehouse_id == get_wh_id(current_user))
    # If specific bill IDs are provided, filter by them
    if bill_ids:
        ids = [int(x.strip()) for x in bill_ids.split(",") if x.strip().isdigit()]
        if ids:
            query = query.where(PayableBill.id.in_(ids))
    else:
        # Otherwise apply status filter (default: pending + partially_paid)
        if status:
            query = query.where(PayableBill.status == status)
        else:
            query = query.where(PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value, PayableStatus.OVERDUE.value]))
    if supplier_id:
        query = query.where(PayableBill.supplier_id == supplier_id)
    if start_date:
        query = query.where(PayableBill.due_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.where(PayableBill.due_date <= datetime.fromisoformat(end_date))
    bills = (await db.execute(query.order_by(PayableBill.due_date))).scalars().all()
    sids = {b.supplier_id for b in bills}
    smap = {}
    if sids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap = {s.id: {"name": s.name, "bank": s.contact_info} for s in sups}
    # 自动标记逾期（due_date < now 且未付）
    from sqlalchemy import update as sql_update
    import sqlalchemy
    expire_sql = (
        sql_update(PayableBill)
        .where(PayableBill.due_date < datetime.now())
        .where(PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value]))
        .values(status=PayableStatus.OVERDUE.value)
    )
    try:
        await db.execute(expire_sql)
        await db.flush()
    except: pass

    return {"data": [{"bill_id": b.id, "supplier": smap.get(b.supplier_id, {}).get("name", ""), "amount": b.amount, "paid_amount": b.paid_amount, "unpaid": b.amount - b.paid_amount, "currency": b.currency, "due_date": str(b.due_date), "bill_number": b.bill_number, "status": b.status} for b in bills]}

@router.get("/supplier-statement")
async def supplier_statement(supplier_id: int = None, current_user: User = Depends(get_current_user),
                              db: AsyncSession = Depends(get_db)):
    """生成供应商对账单"""
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PayableBill)
    if supplier_id: query = query.where(PayableBill.supplier_id == supplier_id)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayableBill.warehouse_id == get_wh_id(current_user))
    result = await db.execute(query.order_by(PayableBill.due_date))
    bills = result.scalars().all()
    sids={b.supplier_id for b in bills}
    smap={}
    if sids:
        sups=(await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
        smap={s.id:{"name":s.name,"contact":s.contact_info} for s in sups}
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb=Workbook(); ws=wb.active; ws.title="供应商对账单"
    hf=PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid"); hfont=Font(bold=True,color="FFFFFF")
    for c,h in enumerate(["供应商","账单号","账单日期","到期日","金额","币种","已付","待付"],1):
        cell=ws.cell(row=1,column=c,value=h); cell.font=hfont; cell.fill=hf
    for r,b in enumerate(bills,2):
        vals=[smap.get(b.supplier_id,{}).get("name",""),b.bill_number,str(b.bill_date)[:10],str(b.due_date)[:10],b.amount,b.currency,b.paid_amount,b.amount-b.paid_amount]
        for c,v in enumerate(vals,1): ws.cell(row=r,column=c,value=str(v) if v is not None else "")
    import io; output=io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=supplier_statement.xlsx"})


# === Plan Templates ===
@router.get("/plan-templates")
async def list_templates(current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(PlanTemplate)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PlanTemplate.warehouse_id == get_wh_id(current_user))
    result = await db.execute(query.order_by(PlanTemplate.created_at.desc()))
    return {"data": [{"id": t.id, "name": t.name, "bill_ids": t.bill_ids} for t in result.scalars().all()]}

@router.delete("/plan-templates/{template_id}")
async def delete_template(template_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(PlanTemplate).where(PlanTemplate.id == template_id))
    t = result.scalar_one_or_none()
    if not t: raise HTTPException(404, "模板不存在")
    if current_user.role != Role.SUPER_ADMIN and t.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "无权限")
    await db.delete(t); await db.flush()
    return {"message": "模板已删除"}

# === Timeline ===
@router.get("/timeline")
async def payable_timeline(current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from datetime import date, timedelta
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    weeks = []
    for w in range(4):
        start = monday + timedelta(weeks=w)
        end = start + timedelta(days=6)
        label = "本周" if w == 0 else "下周" if w == 1 else f"第{w+1}周"
        q = select(func.coalesce(func.sum(PayableBill.amount - PayableBill.paid_amount), 0)).where(
            PayableBill.due_date >= start, PayableBill.due_date <= end,
            PayableBill.status.in_([PayableStatus.PENDING.value, PayableStatus.PARTIALLY_PAID.value])
        )
        if current_user.role != Role.SUPER_ADMIN:
            q = q.where(PayableBill.warehouse_id == get_wh_id(current_user))
        total = float((await db.execute(q)).scalar() or 0)
        weeks.append({"label": label, "start": start.isoformat(), "end": end.isoformat(), "total": total})
    return {"data": weeks}

# === Execute Plan ===
@router.put("/plans/{plan_id}/execute")
async def execute_plan(plan_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    result = await db.execute(select(PayablePlan).where(PayablePlan.id == plan_id))
    p = result.scalar_one_or_none()
    if not p: raise HTTPException(404, "计划不存在")
    if current_user.role != Role.SUPER_ADMIN and p.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "无权限")
    if p.status != PlanStatus.PENDING.value:
        raise HTTPException(400, "该计划已执行或已取消")
    if p.bill_ids:
        bills = (await db.execute(select(PayableBill).where(PayableBill.id.in_(p.bill_ids)))).scalars().all()
        for b in bills:
            if b.status not in (PayableStatus.PAID.value,):
                b.paid_amount = b.amount
                b.paid_at = datetime.now()
                b.payment_method = "银行转账"
                b.status = PayableStatus.PAID.value
    p.status = PlanStatus.EXECUTED.value
    await db.flush()
    return {"message": "计划执行成功，关联账单已自动付款"}

# === Plan Detail ===
@router.get("/plans/{plan_id}/detail")
async def plan_detail(plan_id: int, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(PayablePlan).where(PayablePlan.id == plan_id))
    p = result.scalar_one_or_none()
    if not p: raise HTTPException(404, "计划不存在")
    bills_data = []
    if p.bill_ids:
        bills = (await db.execute(select(PayableBill).where(PayableBill.id.in_(p.bill_ids)))).scalars().all()
        sids = {b.supplier_id for b in bills}
        smap = {}
        if sids:
            sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sids)))).scalars().all()
            smap = {s.id: s.name for s in sups}
        bills_data = [{
            "id": b.id, "supplier_name": smap.get(b.supplier_id, ""),
            "bill_number": b.bill_number, "amount": b.amount,
            "due_date": b.due_date.isoformat() if b.due_date else None,
            "status": b.status, "paid_amount": b.paid_amount,
        } for b in bills]
    return {
        "plan": {"id": p.id, "plan_name": p.plan_name, "planned_date": p.planned_date.isoformat() if p.planned_date else None, "total_amount": p.total_amount, "status": p.status, "remark": p.remark},
        "bills": bills_data,
        "bill_count": len(bills_data),
    }

# === Export Plan ===
@router.get("/plans/export")
async def export_plans(plan_id: int = None, current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    query = select(PayablePlan)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(PayablePlan.warehouse_id == get_wh_id(current_user))
    if plan_id:
        query = query.where(PayablePlan.id == plan_id)
    plans = (await db.execute(query)).scalars().all()

    wb = Workbook(); ws = wb.active; ws.title = "付款计划导出"
    hf = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(["计划名称", "计划日期", "总金额", "状态", "关联账单数"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hfont; cell.fill = hf
    for r, p in enumerate(plans, 2):
        vals = [p.plan_name, str(p.planned_date)[:10] if p.planned_date else "", p.total_amount, p.status, len(p.bill_ids or [])]
        for c, v in enumerate(vals, 1): ws.cell(row=r, column=c, value=str(v) if v is not None else "")
    import io; output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=payment_plans.xlsx"})


# === Monthly Order Volume ===
class MonthOrderReq(BaseModel):
    month: str; order_count: int

@router.get("/monthly-order")
async def get_monthly_order(
    start_month: str = None, end_month: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(MonthlyOrderVolume).where(MonthlyOrderVolume.warehouse_id == get_wh_id(current_user))
    if start_month:
        query = query.where(MonthlyOrderVolume.month >= start_month)
    if end_month:
        query = query.where(MonthlyOrderVolume.month <= end_month)
    result = await db.execute(query.order_by(MonthlyOrderVolume.month))
    return {"data": [{"month": r.month, "order_count": r.order_count} for r in result.scalars().all()]}

@router.post("/monthly-order")
async def save_monthly_order(
    req: MonthOrderReq, current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    # Upsert
    existing = (await db.execute(
        select(MonthlyOrderVolume).where(
            MonthlyOrderVolume.warehouse_id == get_wh_id(current_user),
            MonthlyOrderVolume.month == req.month,
        )
    )).scalar_one_or_none()
    if existing:
        existing.order_count = req.order_count
        existing.updated_by = current_user.id
    else:
        r = MonthlyOrderVolume(
            warehouse_id=get_wh_id(current_user), month=req.month,
            order_count=req.order_count, updated_by=current_user.id,
        )
        db.add(r)
    await db.flush()
    return {"message": "保存成功"}

# === Trend Data ===
@router.get("/trend")
async def payable_trend(
    months: int = 6,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from datetime import date, timedelta
    import calendar

    # Find consumable category id
    cat_result = await db.execute(select(SupplierCategory).where(SupplierCategory.name == "耗材商"))
    cat = cat_result.scalar_one_or_none()
    con_supplier_ids = []
    if cat:
        sup_result = await db.execute(select(Supplier.id).where(Supplier.category_id == cat.id))
        con_supplier_ids = [r[0] for r in sup_result.all()]

    today = date.today()
    months_list = []
    for i in range(months - 1, -1, -1):
        m = today.month - i; y = today.year
        if m <= 0: m += 12; y -= 1
        months_list.append(f"{y}-{m:02d}")

    # Monthly order volumes
    order_q = select(MonthlyOrderVolume).where(
        MonthlyOrderVolume.warehouse_id == get_wh_id(current_user),
        MonthlyOrderVolume.month.in_(months_list),
    )
    order_rows = (await db.execute(order_q)).scalars().all()
    order_map = {r.month: r.order_count for r in order_rows}

    # Monthly consumable spending
    spending_map = {}
    for month in months_list:
        m_start = datetime.fromisoformat(f"{month}-01")
        last_day = calendar.monthrange(int(month[:4]), int(month[5:]))[1]
        m_end = datetime.fromisoformat(f"{month}-{last_day:02d}") + timedelta(days=1) - timedelta(microseconds=1)
        q = select(func.coalesce(func.sum(PayableBill.amount), 0)).where(
            PayableBill.warehouse_id == get_wh_id(current_user),
            PayableBill.bill_date >= m_start,
            PayableBill.bill_date <= m_end,
        )
        if con_supplier_ids:
            q = q.where(PayableBill.supplier_id.in_(con_supplier_ids))
        total = float((await db.execute(q)).scalar() or 0)
        spending_map[month] = total

    # Build response
    data = []
    prev_order = None; prev_spending = None
    warnings = []
    for i, month in enumerate(months_list):
        order = order_map.get(month, 0)
        spending = spending_map.get(month, 0)
        data.append({"month": month, "order_count": order, "consumable_spending": spending})
        # Check anomaly: order flat/down but spending up > 20%
        if i > 0 and prev_order is not None and prev_spending is not None:
            if order <= prev_order and spending > prev_spending * 1.2:
                warnings.append(f"{month}: 订单量持平或下降，但耗材支出上涨超20%")
        prev_order = order; prev_spending = spending

    return {"data": data, "warnings": warnings}
