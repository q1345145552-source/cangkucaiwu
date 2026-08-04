"""系统设置：操作日志 + 数据备份"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text as sa_text
from app.core.timezone import thai_now, thai_today
from datetime import datetime
from app.database import get_db
from app.models.audit_log import AuditLog
from app.models.user import User
from app.core.permissions import get_current_user, Role
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

@router.get("/logs")
async def operation_logs(
    page: int = 1, page_size: int = 20,
    user_id: int = None, action_type: str = None, date: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != Role.SUPER_ADMIN:
        return {"data": [], "total": 0, "message": "仅超级管理员可查看操作日志"}
    query = select(AuditLog); count_q = select(func.count(AuditLog.id))
    if user_id: query = query.where(AuditLog.user_id == user_id); count_q = count_q.where(AuditLog.user_id == user_id)
    if action_type: query = query.where(AuditLog.action_type == action_type); count_q = count_q.where(AuditLog.action_type == action_type)
    if date: query = query.where(func.date(AuditLog.created_at) == date); count_q = count_q.where(func.date(AuditLog.created_at) == date)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(AuditLog.created_at.desc()).offset((page-1)*page_size).limit(page_size))
    logs = result.scalars().all()
    uid_map = {}
    uids = {l.user_id for l in logs if l.user_id}
    if uids:
        users = (await db.execute(select(User).where(User.id.in_(uids)))).scalars().all()
        uid_map = {u.id: u.display_name for u in users}
    return {"data": [{
        "id": l.id, "user_name": uid_map.get(l.user_id, ""),
        "action_type": l.action_type, "module": l.module, "target_id": l.target_id,
        "created_at": l.created_at.isoformat() if l.created_at else None,
    } for l in logs], "total": total, "page": page, "page_size": page_size}

@router.post("/backup")
async def backup_all_data(current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role != Role.SUPER_ADMIN:
        return {"error": "仅超级管理员可操作"}
    wb = Workbook()
    tables = [
        ("users", "id,username,display_name,role,warehouse_id,is_active,created_at"),
        ("warehouses", "id,name,code,is_active,created_at"),
        ("customers", "id,warehouse_id,customer_code,company_name,contact_person,credit_status,credit_limit"),
        ("payment_accounts", "id,warehouse_id,account_name,account_type,account_number,opening_balance"),
        ("suppliers", "id,name,contact_person,contact_info"),
        ("recharge_declarations", "id,warehouse_id,customer_id,declare_date,amount,currency,match_status"),
        ("incoming_flows", "id,warehouse_id,received_date,amount,currency,payer_name,match_status"),
        ("reconciliation_results", "id,warehouse_id,reconciliation_month,match_status,amount_diff"),
        ("income_records", "id,warehouse_id,amount,currency,income_date"),
        ("expense_records", "id,warehouse_id,amount,currency,expense_date"),
        ("expense_funds", "id,warehouse_id,employee_id,amount,purpose,status,remaining_balance"),
        ("reimbursements", "id,warehouse_id,employee_id,total_amount,currency,status"),
        ("payable_bills", "id,warehouse_id,supplier_id,bill_number,due_date,amount,paid_amount,status"),
        ("credit_customers", "id,warehouse_id,customer_id,credit_limit,current_debt,overdue_days,status"),
        ("market_items", "id,warehouse_id,name,quantity,price,status"),
        ("group_orders", "id,warehouse_id,item_name,target_quantity,target_price,status"),
    ]
    hf = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for tname, cols in tables:
        ws = wb.create_sheet(title=tname[:31])
        col_list = cols.split(",")
        for c, col in enumerate(col_list, 1):
            cell = ws.cell(row=1, column=c, value=col.strip())
            cell.font = hfont; cell.fill = hf
        try:
            result = await db.execute(sa_text(f"SELECT {cols} FROM {tname} LIMIT 5000"))
            rows = result.all()
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=str(val)[:1000] if val is not None else "")
        except Exception:
            pass
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename=backup_{thai_now().strftime('%Y%m%d')}.xlsx"})
