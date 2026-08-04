from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.core.timezone import thai_now, thai_today
from datetime import datetime
from app.database import get_db
from app.models.income_expense import IncomeRecord, ExpenseRecord, IncomeExpenseCategory, IncomeExpenseType, CategoryStatus
from app.models.customer import Customer, PaymentAccount
from app.models.supplier import Supplier
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, get_wh_ids, Role, check_staff_permission
from app.schemas.business import IncomeRecordCreate, ExpenseRecordCreate, CategoryCreate

OPERATING_NAMES = ["仓储费", "操作费", "增值服务费", "工资", "电费", "网费", "房租", "耗材", "物流运费", "快递费", "保险费", "税费"]

router = APIRouter()


# ==== Categories ====
@router.get("/categories")
async def list_categories(type: str = None, category_group: str = None, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(IncomeExpenseCategory)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(IncomeExpenseCategory.warehouse_id.in_(get_wh_ids(current_user)))
    if type:
        query = query.where(IncomeExpenseCategory.type == type)
    if category_group:
        query = query.where(IncomeExpenseCategory.category_group == category_group)
    result = await db.execute(query.order_by(IncomeExpenseCategory.sort_order))
    cats = result.scalars().all()
    return {"data": [{"id": c.id, "warehouse_id": c.warehouse_id, "type": c.type,
                      "name": c.name, "sort_order": c.sort_order, "status": c.status, "category_group": c.category_group} for c in cats]}

@router.post("/categories")
async def create_category(req: CategoryCreate, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    # 自动归类：运营类 vs 其他
    from app.api.v1.income_expense import OPERATING_NAMES
    category_group = "operating" if req.name in OPERATING_NAMES else "other"
    c = IncomeExpenseCategory(warehouse_id=get_wh_id(current_user), category_group=category_group, **req.model_dump())
    db.add(c); await db.flush(); return {"id": c.id, "message": "创建成功"}

# ==== Income ====
@router.get("/income")
async def list_income(
    page: int = 1, page_size: int = 20, month: str = None,
    account_id: int = None, customer_id: int = None,
    category_id: int = None,
    category_group: str = None,
    currency: str = None,
    search: str = None,
    current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(IncomeRecord); count_q = select(func.count(IncomeRecord.id))
    if category_group:
        query = query.join(IncomeExpenseCategory, IncomeRecord.category_id == IncomeExpenseCategory.id)
        query = query.where(IncomeExpenseCategory.category_group == category_group)
        count_q = count_q.join(IncomeExpenseCategory, IncomeRecord.category_id == IncomeExpenseCategory.id)
        count_q = count_q.where(IncomeExpenseCategory.category_group == category_group)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(IncomeRecord.warehouse_id.in_(get_wh_ids(current_user)))
        count_q = count_q.where(IncomeRecord.warehouse_id.in_(get_wh_ids(current_user)))
    if month:
        query = query.where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == month)
        count_q = count_q.where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == month)
    if account_id:
        query = query.where(IncomeRecord.account_id == account_id)
        count_q = count_q.where(IncomeRecord.account_id == account_id)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(IncomeRecord.income_date.desc()).offset((page-1)*page_size).limit(page_size))
    records = result.scalars().all()

    cat_ids = {r.category_id for r in records}
    acc_ids = {r.account_id for r in records}
    cust_ids = {r.customer_id for r in records if r.customer_id}
    cat_map = {}; acc_map = {}; cust_map = {}
    if cat_ids:
        cats = (await db.execute(select(IncomeExpenseCategory).where(IncomeExpenseCategory.id.in_(cat_ids)))).scalars().all()
        cat_map = {c.id: c.name for c in cats}
    if acc_ids:
        accs = (await db.execute(select(PaymentAccount).where(PaymentAccount.id.in_(acc_ids)))).scalars().all()
        acc_map = {a.id: a.account_name for a in accs}
    if cust_ids:
        custs = (await db.execute(select(Customer).where(Customer.id.in_(cust_ids)))).scalars().all()
        cust_map = {c.id: c.company_name for c in custs}

    return {"data": [{
        "id": r.id, "warehouse_id": r.warehouse_id, "category_id": r.category_id,
        "account_id": r.account_id, "customer_id": r.customer_id,
        "amount": r.amount, "currency": r.currency,
        "income_date": r.income_date.isoformat() if r.income_date else None,
        "voucher": r.voucher, "remark": r.remark,
        "category_name": cat_map.get(r.category_id, ""),
        "account_name": acc_map.get(r.account_id, ""),
        "customer_name": cust_map.get(r.customer_id, ""),
    } for r in records], "total": total, "page": page, "page_size": page_size}

@router.post("/income")
async def create_income(req: IncomeRecordCreate, current_user: User = Depends(get_current_user),
                        db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role == Role.STAFF and "收付款管理" not in (current_user.extra_permissions or []):
        raise HTTPException(403, "无确认入账权限")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    r = IncomeRecord(
        warehouse_id=get_wh_id(current_user), category_id=req.category_id,
        account_id=req.account_id, customer_id=req.customer_id,
        amount=req.amount, currency=req.currency,
        income_date=datetime.fromisoformat(req.income_date),
        voucher=req.voucher, remark=req.remark, confirmed_by=current_user.id,
    )
    db.add(r); await db.flush(); return {"id": r.id, "message": "收款记录创建成功"}

# ==== Income Edit / Delete ====
@router.put("/income/{income_id}")
async def update_income(income_id: int, req: IncomeRecordCreate, current_user: User = Depends(get_current_user),
                        db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    r = (await db.execute(select(IncomeRecord).where(IncomeRecord.id == income_id))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "记录不存在")
    if current_user.role != Role.SUPER_ADMIN and r.warehouse_id not in get_wh_ids(current_user):
        raise HTTPException(403, "无权限")
    r.category_id = req.category_id
    r.account_id = req.account_id
    r.amount = req.amount
    r.currency = req.currency
    r.income_date = datetime.fromisoformat(req.income_date)
    r.remark = req.remark
    await db.flush()
    return {"message": "更新成功"}

@router.delete("/income/{income_id}")
async def delete_income(income_id: int, current_user: User = Depends(get_current_user),
                        db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    r = (await db.execute(select(IncomeRecord).where(IncomeRecord.id == income_id))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "记录不存在")
    if current_user.role != Role.SUPER_ADMIN and r.warehouse_id not in get_wh_ids(current_user):
        raise HTTPException(403, "无权限")
    await db.delete(r)
    await db.flush()
    return {"message": "删除成功"}


# ==== Expense ====
@router.get("/expense")
async def list_expense(
    page: int = 1, page_size: int = 20, month: str = None,
    category_id: int = None, account_id: int = None,
    start_date: str = None, end_date: str = None,
    category_group: str = None,
    currency: str = None,
    current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(ExpenseRecord); count_q = select(func.count(ExpenseRecord.id))
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(ExpenseRecord.warehouse_id.in_(get_wh_ids(current_user)))
        count_q = count_q.where(ExpenseRecord.warehouse_id.in_(get_wh_ids(current_user)))
    if month:
        query = query.where(func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == month)
        count_q = count_q.where(func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == month)
    if category_id:
        query = query.where(ExpenseRecord.category_id == category_id)
        count_q = count_q.where(ExpenseRecord.category_id == category_id)
    if account_id:
        query = query.where(ExpenseRecord.account_id == account_id)
        count_q = count_q.where(ExpenseRecord.account_id == account_id)
    if start_date:
        query = query.where(ExpenseRecord.expense_date >= datetime.strptime(start_date, "%Y-%m-%d"))
        count_q = count_q.where(ExpenseRecord.expense_date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.where(ExpenseRecord.expense_date <= datetime.strptime(end_date, "%Y-%m-%d"))
        count_q = count_q.where(ExpenseRecord.expense_date <= datetime.strptime(end_date, "%Y-%m-%d"))
    if currency:
        query = query.where(ExpenseRecord.currency == currency)
        count_q = count_q.where(ExpenseRecord.currency == currency)
    if category_group:
        query = query.join(IncomeExpenseCategory, ExpenseRecord.category_id == IncomeExpenseCategory.id)
        query = query.where(IncomeExpenseCategory.category_group == category_group)
        count_q = count_q.join(IncomeExpenseCategory, ExpenseRecord.category_id == IncomeExpenseCategory.id)
        count_q = count_q.where(IncomeExpenseCategory.category_group == category_group)
    total = (await db.execute(count_q)).scalar()
    result = await db.execute(query.order_by(ExpenseRecord.expense_date.desc()).offset((page-1)*page_size).limit(page_size))
    records = result.scalars().all()

    cat_ids = {r.category_id for r in records}
    acc_ids = {r.account_id for r in records}
    sup_ids = {r.supplier_id for r in records if r.supplier_id}
    cat_map = {}; acc_map = {}; sup_map = {}
    if cat_ids:
        cats = (await db.execute(select(IncomeExpenseCategory).where(IncomeExpenseCategory.id.in_(cat_ids)))).scalars().all()
        cat_map = {c.id: c.name for c in cats}
    if acc_ids:
        accs = (await db.execute(select(PaymentAccount).where(PaymentAccount.id.in_(acc_ids)))).scalars().all()
        acc_map = {a.id: a.account_name for a in accs}
    if sup_ids:
        sups = (await db.execute(select(Supplier).where(Supplier.id.in_(sup_ids)))).scalars().all()
        sup_map = {s.id: s.name for s in sups}

    return {"data": [{
        "id": r.id, "warehouse_id": r.warehouse_id, "category_id": r.category_id,
        "account_id": r.account_id, "supplier_id": r.supplier_id,
        "amount": r.amount, "currency": r.currency,
        "expense_date": r.expense_date.isoformat() if r.expense_date else None,
        "voucher": r.voucher, "remark": r.remark,
        "category_name": cat_map.get(r.category_id, ""),
        "account_name": acc_map.get(r.account_id, ""),
        "supplier_name": sup_map.get(r.supplier_id, ""),
    } for r in records], "total": total, "page": page, "page_size": page_size}

@router.post("/expense")
async def create_expense(req: ExpenseRecordCreate, current_user: User = Depends(get_current_user),
                         db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role == Role.STAFF and "收付款管理" not in (current_user.extra_permissions or []):
        raise HTTPException(403, "无确认出账权限")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    r = ExpenseRecord(
        warehouse_id=get_wh_id(current_user), category_id=req.category_id,
        account_id=req.account_id, supplier_id=req.supplier_id,
        amount=req.amount, currency=req.currency,
        expense_date=datetime.fromisoformat(req.expense_date),
        voucher=req.voucher, remark=req.remark, approved_by=current_user.id,
    )
    db.add(r); await db.flush(); return {"id": r.id, "message": "付款记录创建成功"}

# ==== Ledger ====
@router.get("/ledger")
async def ledger(
    page: int = 1, page_size: int = 30,
    month: str = None,
    source: str = None,
    flow_type: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    wh_id = get_wh_id(current_user)
    today = __import__('datetime').thai_today()
    cur_month = f"{today.year}-{today.month:02d}"

    # Build SQL UNION ALL for pagination at DB level
    # Each subquery has: date_val, amount, currency, flow_type, source, source_label, remark, ref_no, raw_id
    from sqlalchemy import text

    # NOTE: alias/col are internal constants (never user input); wh_id and month
    # are passed as bind parameters to avoid SQL injection.
    def wh_cond(alias: str):
        return f"{alias}.warehouse_id = :wh_id"

    def month_cond(alias: str, col: str):
        return f"to_char({alias}.{col}, 'YYYY-MM') = :month" if month else "1=1"

    unions = []
    params = {}

    # 1. 充值申报
    if not source or source == "recharge":
        unions.append(f"""
            SELECT d.declare_date AS date_val, d.amount, d.currency, 'income' AS flow_type, 'recharge' AS source,
                   '充值申报' AS source_label,
                   COALESCE('客户: ' || cust.company_name, '充值') AS remark,
                   'RC-' || d.id AS ref_no, d.id AS raw_id, 1 AS sort_ord
            FROM recharge_declarations d
            LEFT JOIN customers cust ON d.customer_id = cust.id
            WHERE {wh_cond('d')} AND {month_cond('d', 'declare_date')}
        """)

    # 2. 到账流水
    if not source or source == "incoming":
        unions.append(f"""
            SELECT f.received_date AS date_val, f.amount, f.currency, 'income' AS flow_type, 'incoming' AS source,
                   '到账流水' AS source_label,
                   '付款方: ' || COALESCE(f.payer_name, '') || COALESCE(' (' || f.payment_method || ')', '') AS remark,
                   'FL-' || f.id AS ref_no, f.id AS raw_id, 2 AS sort_ord
            FROM incoming_flows f
            WHERE {wh_cond('f')} AND {month_cond('f', 'received_date')}
        """)

    # 3. 备用金领用
    if not source or source == "expense_fund":
        unions.append(f"""
            SELECT ef.receive_date AS date_val, ef.amount, 'THB' AS currency, 'expense' AS flow_type,
                   'expense_fund' AS source, '备用金领用' AS source_label,
                   '领用人: ' || COALESCE(u.display_name, '') || ' (' || COALESCE(ef.purpose, '') || ')' AS remark,
                   'EF-' || ef.id AS ref_no, ef.id AS raw_id, 3 AS sort_ord
            FROM expense_funds ef
            LEFT JOIN users u ON ef.employee_id = u.id
            WHERE {wh_cond('ef')} AND ef.amount > 0 AND {month_cond('ef', 'receive_date')}
        """)

    # 4. 备用金开销
    if not source or source == "fund_item":
        unions.append(f"""
            SELECT efi.expense_date AS date_val, efi.amount, COALESCE(efi.currency, 'THB') AS currency,
                   'expense' AS flow_type, 'fund_item' AS source, '备用金开销' AS source_label,
                   COALESCE(u2.display_name, '') || ': ' || efi.category || ' - ' || COALESCE(efi.description, '') AS remark,
                   'FI-' || efi.id AS ref_no, efi.id AS raw_id, 4 AS sort_ord
            FROM expense_fund_items efi
            JOIN expense_funds ef2 ON efi.fund_id = ef2.id
            LEFT JOIN users u2 ON ef2.employee_id = u2.id
            WHERE {wh_cond('ef2')} AND {month_cond('efi', 'expense_date')}
        """)

    # 5. 报销出款
    if not source or source == "reimbursement":
        unions.append(f"""
            SELECT COALESCE(rb.paid_at, rb.submit_date) AS date_val, rb.total_amount AS amount,
                   COALESCE(rb.currency, 'THB') AS currency, 'expense' AS flow_type,
                   'reimbursement' AS source, '报销' AS source_label,
                   '报销人: ' || COALESCE(u3.display_name, '') AS remark,
                   'RB-' || rb.id AS ref_no, rb.id AS raw_id, 5 AS sort_ord
            FROM reimbursements rb
            LEFT JOIN users u3 ON rb.employee_id = u3.id
            WHERE rb.status = 'paid' AND {wh_cond('rb')} AND {month_cond('rb', 'paid_at')}
        """)

    # 6. 应付账款付款
    if not source or source == "payable":
        unions.append(f"""
            SELECT COALESCE(pb.paid_at, pb.bill_date) AS date_val, pb.paid_amount AS amount,
                   COALESCE(pb.currency, 'THB') AS currency, 'expense' AS flow_type,
                   'payable' AS source, '应付账款' AS source_label,
                   '账单: ' || COALESCE(pb.bill_number, '') || COALESCE(' (' || pb.remark || ')', '') AS remark,
                   'PB-' || pb.id AS ref_no, pb.id AS raw_id, 6 AS sort_ord
            FROM payable_bills pb
            WHERE pb.paid_amount > 0 AND {wh_cond('pb')} AND {month_cond('pb', 'paid_at')}
        """)

    # 7. 手工收支
    if not source or source == "manual":
        unions.append(f"""
            SELECT ir.income_date AS date_val, ir.amount, COALESCE(ir.currency, 'THB') AS currency,
                   'income' AS flow_type, 'manual' AS source, '手工收支' AS source_label,
                   COALESCE(iec.name, '') || COALESCE(' - ' || ir.remark, '') AS remark,
                   'IN-' || ir.id AS ref_no, ir.id AS raw_id, 7 AS sort_ord
            FROM income_records ir
            LEFT JOIN income_expense_categories iec ON ir.category_id = iec.id
            WHERE {wh_cond('ir')} AND {month_cond('ir', 'income_date')}
        """)
        unions.append(f"""
            SELECT er.expense_date AS date_val, er.amount, COALESCE(er.currency, 'THB') AS currency,
                   'expense' AS flow_type, 'manual' AS source, '手工收支' AS source_label,
                   COALESCE(iec2.name, '') || COALESCE(' - ' || er.remark, '') AS remark,
                   'EX-' || er.id AS ref_no, er.id AS raw_id, 8 AS sort_ord
            FROM expense_records er
            LEFT JOIN income_expense_categories iec2 ON er.category_id = iec2.id
            WHERE {wh_cond('er')} AND {month_cond('er', 'expense_date')}
        """)

    if not unions:
        return {"data": [], "total": 0, "page": page, "page_size": page_size,
                "total_income": 0, "total_expense": 0, "net": 0,
                "card_income": 0, "card_expense": 0, "card_net": 0}

    full_union = " UNION ALL ".join(unions)

    # Build the wrapped query
    type_filter = ""
    if flow_type == "income":
        type_filter = "WHERE flow_type = 'income'"
    elif flow_type == "expense":
        type_filter = "WHERE flow_type = 'expense'"

    sql = f"""
        WITH all_flows AS (
            {full_union}
        )
        SELECT date_val, amount, currency, flow_type, source, source_label, remark, ref_no, raw_id
        FROM all_flows
        {type_filter}
        ORDER BY date_val DESC NULLS LAST
        LIMIT :page_size OFFSET :offset
    """

    count_sql = f"""
        WITH all_flows AS (
            {full_union}
        )
        SELECT COUNT(*) AS cnt
        FROM all_flows
        {type_filter}
    """

    # Bind params shared by every statement (all embed full_union)
    base_params = {"wh_id": wh_id}
    if month:
        base_params["month"] = month

    offset = (page - 1) * page_size
    result = await db.execute(text(sql), {**base_params, "page_size": page_size, "offset": offset})
    rows = result.fetchall()

    total_row = (await db.execute(text(count_sql), base_params)).fetchone()
    total_all = total_row[0] if total_row else 0

    # Compute summary: income/expense totals for the filtered data
    summary_sql = f"""
        WITH all_flows AS (
            {full_union}
        )
        SELECT
            COALESCE(SUM(CASE WHEN flow_type = 'income' THEN amount ELSE 0 END), 0) AS total_income,
            COALESCE(SUM(CASE WHEN flow_type = 'expense' THEN amount ELSE 0 END), 0) AS total_expense
        FROM all_flows
        {type_filter}
    """
    sum_result = await db.execute(text(summary_sql), base_params)
    sum_row = sum_result.fetchone()
    total_income = float(sum_row[0]) if sum_row else 0
    total_expense = float(sum_row[1]) if sum_row else 0

    # Card values: current month only
    card_sql = f"""
        WITH all_flows AS (
            {full_union}
        )
        SELECT
            COALESCE(SUM(CASE WHEN flow_type = 'income' AND to_char(date_val, 'YYYY-MM') = :cur_month THEN amount ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN flow_type = 'expense' AND to_char(date_val, 'YYYY-MM') = :cur_month THEN amount ELSE 0 END), 0)
        FROM all_flows
        {type_filter}
    """
    card_result = await db.execute(text(card_sql), {**base_params, "cur_month": cur_month})
    card_row = card_result.fetchone()
    card_income = float(card_row[0]) if card_row else 0
    card_expense = float(card_row[1]) if card_row else 0

    # Build response
    data = []
    for row in rows:
        data.append({
            "date": str(row[0])[:10] if row[0] else "",
            "amount": float(row[1]) if row[1] else 0,
            "currency": row[2] or "THB",
            "type": row[3],
            "source": row[4],
            "source_label": row[5] or "",
            "remark": row[6] or "",
            "ref_no": row[7] or "",
            "id": row[8] or 0,
        })

    return {
        "data": data, "total": total_all, "page": page, "page_size": page_size,
        "total_income": total_income, "total_expense": total_expense,
        "net": total_income - total_expense,
        "card_income": card_income, "card_expense": card_expense,
        "card_net": card_income - card_expense,
    }

# ==== Ledger Export ====
@router.get("/ledger/export")
async def ledger_export(
    month: str = None, source: str = None, flow_type: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    # Reuse same aggregation logic
    r = await ledger(page=1, page_size=99999, month=month, source=source, flow_type=flow_type, current_user=current_user, db=db)
    data = r.get("data", [])

    wb = __import__('openpyxl').Workbook()
    ws = wb.active; ws.title = "资金流水"
    hf = __import__('openpyxl').styles.PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = __import__('openpyxl').styles.Font(bold=True, color="FFFFFF")
    headers = ["日期", "金额", "币种", "类型", "来源模块", "说明", "关联单号"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = hfont; cell.fill = hf
    for i, row in enumerate(data, 2):
        type_label = "收入" if row.get("type") == "income" else "支出"
        vals = [
            row.get("date", ""), row.get("amount", 0), row.get("currency", ""),
            type_label, row.get("source_label", ""), row.get("remark", ""), row.get("ref_no", ""),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=str(v) if v is not None else "")
    output = __import__('io').BytesIO(); wb.save(output); output.seek(0)
    filename = f"ledger_{month or 'all'}.xlsx"
    return __import__('fastapi').responses.StreamingResponse(
        output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==== Income Export ====
@router.get("/income/export")
async def export_income(
    month: str = None,
    category_group: str = None,
    currency: str = None,
    search: str = None,
    category_id: int = None,
    start_date: str = None,
    end_date: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(IncomeRecord)
    if current_user.role != Role.SUPER_ADMIN:
        query = query.where(IncomeRecord.warehouse_id.in_(get_wh_ids(current_user)))
    if month:
        query = query.where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == month)
    if category_group:
        query = query.join(IncomeExpenseCategory, IncomeRecord.category_id == IncomeExpenseCategory.id)
        query = query.where(IncomeExpenseCategory.category_group == category_group)
    if currency:
        query = query.where(IncomeRecord.currency == currency)
    if search:
        query = query.where(IncomeRecord.remark.ilike(f"%{search}%"))
    if category_id:
        query = query.where(IncomeRecord.category_id == category_id)
    if start_date:
        query = query.where(IncomeRecord.income_date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.where(IncomeRecord.income_date <= datetime.strptime(end_date, "%Y-%m-%d"))
    result = await db.execute(query.order_by(IncomeRecord.income_date.desc()))
    records = result.scalars().all()

    cat_ids = {r.category_id for r in records}
    acc_ids = {r.account_id for r in records}
    cat_map = {}; acc_map = {}
    if cat_ids:
        cats = (await db.execute(select(IncomeExpenseCategory).where(IncomeExpenseCategory.id.in_(cat_ids)))).scalars().all()
        cat_map = {c.id: c.name for c in cats}
    if acc_ids:
        accs = (await db.execute(select(PaymentAccount).where(PaymentAccount.id.in_(acc_ids)))).scalars().all()
        acc_map = {a.id: a.account_name for a in accs}

    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "其他收入"
    ws.append(["日期", "分类", "物品说明", "金额", "币种", "账户"])
    for r in records:
        ws.append([
            r.income_date.strftime("%Y-%m-%d") if r.income_date else "",
            cat_map.get(r.category_id, ""),
            r.remark or "",
            r.amount,
            r.currency,
            acc_map.get(r.account_id, ""),
        ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": "attachment; filename=other_income.xlsx"})

# ==== Monthly Summary ====
@router.get("/monthly-summary")
async def monthly_summary(month: str, category_group: str = None, current_user: User = Depends(get_current_user),
                          db: AsyncSession = Depends(get_db)):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    iq = select(func.sum(IncomeRecord.amount)).where(func.to_char(IncomeRecord.income_date, 'YYYY-MM') == month)
    eq = select(func.sum(ExpenseRecord.amount)).where(func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == month)
    if category_group:
        iq = iq.join(IncomeExpenseCategory, IncomeRecord.category_id == IncomeExpenseCategory.id).where(IncomeExpenseCategory.category_group == category_group)
        eq = eq.join(IncomeExpenseCategory, ExpenseRecord.category_id == IncomeExpenseCategory.id).where(IncomeExpenseCategory.category_group == category_group)
    if current_user.role != Role.SUPER_ADMIN:
        iq = iq.where(IncomeRecord.warehouse_id.in_(get_wh_ids(current_user)))
        eq = eq.where(ExpenseRecord.warehouse_id.in_(get_wh_ids(current_user)))
    ti = (await db.execute(iq)).scalar() or 0
    te = (await db.execute(eq)).scalar() or 0
    return {"month": month, "total_income": float(ti), "total_expense": float(te), "net": float(ti - te)}
# ==== Operating Dashboard ====
@router.get("/operating-dashboard")
async def operating_dashboard(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    from datetime import date, timedelta
    import calendar

    today = thai_today()
    months_list = []
    for i in range(11, -1, -1):
        m = today.month - i; y = today.year
        if m <= 0: m += 12; y -= 1
        months_list.append(f"{y}-{m:02d}")

    # Monthly recharge totals (income)
    from app.models.recharge import RechargeDeclaration
    recharge_map = {}
    for month in months_list:
        m_start = f"{month}-01"
        last_day = calendar.monthrange(int(month[:4]), int(month[5:]))[1]
        m_end = f"{month}-{last_day:02d}"
        q = select(func.coalesce(func.sum(RechargeDeclaration.amount), 0)).where(
            RechargeDeclaration.warehouse_id.in_(get_wh_ids(current_user)),
            func.to_char(RechargeDeclaration.declare_date, 'YYYY-MM') == month,
        )
        total = float((await db.execute(q)).scalar() or 0)
        recharge_map[month] = total

    # Monthly operating expense totals
    expense_map = {}
    for month in months_list:
        m_start = f"{month}-01"
        last_day = calendar.monthrange(int(month[:4]), int(month[5:]))[1]
        m_end = f"{month}-{last_day:02d}"
        q = select(func.coalesce(func.sum(ExpenseRecord.amount), 0)).where(
            ExpenseRecord.warehouse_id.in_(get_wh_ids(current_user)),
            func.to_char(ExpenseRecord.expense_date, 'YYYY-MM') == month,
        )
        # Join to filter only operating categories
        q = q.join(IncomeExpenseCategory, ExpenseRecord.category_id == IncomeExpenseCategory.id)
        q = q.where(IncomeExpenseCategory.category_group == "operating")
        total = float((await db.execute(q)).scalar() or 0)
        expense_map[month] = total

    data = []
    for month in months_list:
        income = recharge_map.get(month, 0)
        expense = expense_map.get(month, 0)
        data.append({
            "month": month,
            "recharge_income": income,
            "operating_expense": expense,
            "net": income - expense,
        })

    cur_month = f"{today.year}-{today.month:02d}"
    cur_income = recharge_map.get(cur_month, 0)
    cur_expense = expense_map.get(cur_month, 0)

    return {
        "data": data,
        "current_month": cur_month,
        "current_income": cur_income,
        "current_expense": cur_expense,
        "current_net": cur_income - cur_expense,
    }

