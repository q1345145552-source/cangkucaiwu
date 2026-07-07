from fastapi.responses import StreamingResponse
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from datetime import datetime, date, timedelta
from app.database import get_db
from app.models.recharge import RechargeDeclaration, IncomingFlow, ReconciliationResult, ReconMatchStatus, MatchStatus
from app.models.customer import Customer
from app.models.user import User
from app.core.permissions import get_current_user, get_wh_id, Role

router = APIRouter()

@router.get("/unmatched")
async def get_unmatched(
    start_date: str = None, end_date: str = None,
    warehouse_id: int = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    wh_id = warehouse_id or get_wh_id(current_user)
    if current_user.role != Role.SUPER_ADMIN:
        if warehouse_id and warehouse_id != get_wh_id(current_user):
            raise HTTPException(403, "无权查看其他仓库")
        wh_id = get_wh_id(current_user)
    today = date.today()
    sd = datetime.fromisoformat(start_date) if start_date else datetime(today.year, today.month, 1)
    ed = datetime.fromisoformat(end_date) if end_date else datetime(today.year, today.month, today.day) + timedelta(days=1) - timedelta(microseconds=1)

    decl_q = select(RechargeDeclaration).where(and_(
        RechargeDeclaration.warehouse_id == wh_id,
        RechargeDeclaration.declare_date >= sd,
        RechargeDeclaration.declare_date <= ed,
        RechargeDeclaration.match_status == MatchStatus.UNMATCHED,
        RechargeDeclaration.screenshot.isnot(None),
    )).order_by(RechargeDeclaration.declare_date.desc())
    decls = (await db.execute(decl_q)).scalars().all()
    cust_ids = {d.customer_id for d in decls}
    cust_map = {}
    if cust_ids:
        custs = (await db.execute(select(Customer).where(Customer.id.in_(cust_ids)))).scalars().all()
        cust_map = {c.id: c for c in custs}
    declarations = [{
        "id": d.id, "customer_id": d.customer_id,
        "customer_name": cust_map[d.customer_id].company_name if d.customer_id in cust_map else "",
        "customer_code": cust_map[d.customer_id].customer_code if d.customer_id in cust_map else "",
        "declare_date": d.declare_date.isoformat() if d.declare_date else None,
        "amount": d.amount, "currency": d.currency or "THB",
        "screenshot": d.screenshot,
        "payment_method": d.payment_method,
    } for d in decls]

    flow_q = select(IncomingFlow).where(and_(
        IncomingFlow.warehouse_id == wh_id,
        IncomingFlow.received_date >= sd,
        IncomingFlow.received_date <= ed,
        IncomingFlow.match_status == MatchStatus.UNMATCHED,
    )).order_by(IncomingFlow.received_date.desc())
    flows = (await db.execute(flow_q)).scalars().all()
    incoming = [{
        "id": f.id, "payer_name": f.payer_name,
        "received_date": f.received_date.isoformat() if f.received_date else None,
        "amount": f.amount, "currency": f.currency or "THB",
        "payment_method": f.payment_method,
    } for f in flows]
    return {"declarations": declarations, "incoming": incoming}

@router.post("/manual-match")
async def manual_match(
    declaration_id: int = Query(...), flow_id: int = Query(...),
    handling_note: str = Query(""),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    if current_user.role not in (Role.SUPER_ADMIN, Role.WAREHOUSE_ADMIN):
        raise HTTPException(403, "无权限")
    decl = (await db.execute(select(RechargeDeclaration).where(RechargeDeclaration.id == declaration_id))).scalar_one_or_none()
    flow = (await db.execute(select(IncomingFlow).where(IncomingFlow.id == flow_id))).scalar_one_or_none()
    if not decl or not flow: raise HTTPException(404, "记录不存在")
    if decl.match_status != MatchStatus.UNMATCHED or flow.match_status != MatchStatus.UNMATCHED:
        raise HTTPException(400, "所选记录已被匹配")
    if current_user.role != Role.SUPER_ADMIN:
        if decl.warehouse_id != get_wh_id(current_user) or flow.warehouse_id != get_wh_id(current_user):
            raise HTTPException(403, "只能匹配自己仓库的记录")
    rr = ReconciliationResult(
        warehouse_id=decl.warehouse_id,
        reconciliation_month=decl.declare_date.strftime("%Y-%m") if decl.declare_date else datetime.now().strftime("%Y-%m"),
        declaration_id=decl.id, flow_id=flow.id,
        match_status=ReconMatchStatus.MANUAL_MATCHED,
        amount_diff=abs(decl.amount - flow.amount),
        handling_note=handling_note, confirmed_by=current_user.id,
    )
    db.add(rr); decl.match_status = MatchStatus.MATCHED; flow.match_status = MatchStatus.MATCHED
    await db.flush(); return {"message": "匹配成功"}

@router.post("/unmatch")
async def unmatch_record(
    record_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    rr = (await db.execute(select(ReconciliationResult).where(ReconciliationResult.id == record_id))).scalar_one_or_none()
    if not rr: raise HTTPException(404, "记录不存在")
    if current_user.role != Role.SUPER_ADMIN and rr.warehouse_id != get_wh_id(current_user):
        raise HTTPException(403, "无权限")
    if rr.declaration_id:
        d = (await db.execute(select(RechargeDeclaration).where(RechargeDeclaration.id == rr.declaration_id))).scalar_one_or_none()
        if d: d.match_status = MatchStatus.UNMATCHED
    if rr.flow_id:
        f = (await db.execute(select(IncomingFlow).where(IncomingFlow.id == rr.flow_id))).scalar_one_or_none()
        if f: f.match_status = MatchStatus.UNMATCHED
    await db.delete(rr); await db.flush(); return {"message": "已解除匹配"}

@router.get("/results")
async def list_results(
    start_date: str = None, end_date: str = None,
    warehouse_id: int = None,
    search: str = None, search_code: str = None,
    page: int = 1, page_size: int = 50,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(ReconciliationResult); count_q = select(func.count(ReconciliationResult.id))
    if current_user.role != Role.SUPER_ADMIN:
        if warehouse_id and warehouse_id != get_wh_id(current_user):
            raise HTTPException(403, "无权查看其他仓库")
        query = query.where(ReconciliationResult.warehouse_id == get_wh_id(current_user))
        count_q = count_q.where(ReconciliationResult.warehouse_id == get_wh_id(current_user))
    elif warehouse_id:
        query = query.where(ReconciliationResult.warehouse_id == warehouse_id)
        count_q = count_q.where(ReconciliationResult.warehouse_id == warehouse_id)
    if start_date:
        query = query.where(ReconciliationResult.created_at >= datetime.fromisoformat(start_date))
        count_q = count_q.where(ReconciliationResult.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        edt = datetime.fromisoformat(end_date) + timedelta(days=1) - timedelta(microseconds=1)
        query = query.where(ReconciliationResult.created_at <= edt)
        count_q = count_q.where(ReconciliationResult.created_at <= edt)

    # Fetch all matched records for the time range (may need filtering by customer later)
    all_records = (await db.execute(query.order_by(ReconciliationResult.created_at.desc()))).scalars().all()

    # Fetch related data
    decl_ids = [r.declaration_id for r in all_records if r.declaration_id]
    flow_ids = [r.flow_id for r in all_records if r.flow_id]
    decl_map = {}; flow_map = {}; cust_map = {}
    if decl_ids:
        decls = (await db.execute(select(RechargeDeclaration).where(RechargeDeclaration.id.in_(decl_ids)))).scalars().all()
        decl_map = {d.id: d for d in decls}
        cust_ids = {d.customer_id for d in decls}
        if cust_ids:
            custs = (await db.execute(select(Customer).where(Customer.id.in_(cust_ids)))).scalars().all()
            cust_map = {c.id: c for c in custs}
    if flow_ids:
        flows = (await db.execute(select(IncomingFlow).where(IncomingFlow.id.in_(flow_ids)))).scalars().all()
        flow_map = {f.id: f for f in flows}

    # Build result items
    items = []
    for r in all_records:
        d = decl_map.get(r.declaration_id)
        cid = d.customer_id if d else None
        c = cust_map.get(cid)
        customer_name = c.company_name if c else ""
        customer_code = c.customer_code if c else ""
        # Apply search filter
        if search and search not in customer_name:
            continue
        if search_code and search_code not in customer_code:
            continue
        f = flow_map.get(r.flow_id)
        items.append({
            "id": r.id, "warehouse_id": r.warehouse_id,
            "reconciliation_month": r.reconciliation_month,
            "declaration_id": r.declaration_id, "flow_id": r.flow_id,
            "match_status": r.match_status or "",
            "amount_diff": r.amount_diff, "handling_note": r.handling_note,
            "customer_code": customer_code, "customer_name": customer_name,
            "decl_amount": d.amount if d else None,
            "decl_currency": d.currency if d else None,
            "flow_payer": f.payer_name if f else "",
            "flow_amount": f.amount if f else None,
            "flow_currency": f.currency if f else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })

    total = len(items)
    # Compute total matched amounts
    total_matched_decl = sum((it.get("decl_amount") or 0) for it in items)
    total_matched_flow = sum((it.get("flow_amount") or 0) for it in items)
    # Apply pagination after filtering
    start_idx = (page - 1) * page_size
    paged = items[start_idx:start_idx + page_size]
    return {"data": paged, "total": total, "page": page, "page_size": page_size,
            "total_matched_decl": total_matched_decl, "total_matched_flow": total_matched_flow}

@router.get("/export")
async def export_reconciliation(
    start_date: str = None, end_date: str = None,
    warehouse_id: int = None,
    search: str = None, search_code: str = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role == Role.SUPER_ADMIN:
        raise HTTPException(403, "超级管理员请使用各仓库管理员账号操作")
    query = select(ReconciliationResult)
    if current_user.role != Role.SUPER_ADMIN:
        if warehouse_id and warehouse_id != get_wh_id(current_user):
            raise HTTPException(403, "无权导出其他仓库")
        query = query.where(ReconciliationResult.warehouse_id == get_wh_id(current_user))
    elif warehouse_id:
        query = query.where(ReconciliationResult.warehouse_id == warehouse_id)
    if start_date:
        query = query.where(ReconciliationResult.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        edt = datetime.fromisoformat(end_date) + timedelta(days=1) - timedelta(microseconds=1)
        query = query.where(ReconciliationResult.created_at <= edt)

    records = (await db.execute(query.order_by(ReconciliationResult.created_at.desc()))).scalars().all()
    decl_ids = [r.declaration_id for r in records if r.declaration_id]
    flow_ids = [r.flow_id for r in records if r.flow_id]
    decl_map = {}; flow_map = {}; cust_map = {}
    if decl_ids:
        decls = (await db.execute(select(RechargeDeclaration).where(RechargeDeclaration.id.in_(decl_ids)))).scalars().all()
        decl_map = {d.id: d for d in decls}
        cust_ids = {d.customer_id for d in decls}
        if cust_ids:
            custs = (await db.execute(select(Customer).where(Customer.id.in_(cust_ids)))).scalars().all()
            cust_map = {c.id: c for c in custs}
    if flow_ids:
        flows = (await db.execute(select(IncomingFlow).where(IncomingFlow.id.in_(flow_ids)))).scalars().all()
        flow_map = {f.id: f for f in flows}

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "对账记录"
    hf = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    headers = ["匹配时间", "客户编号", "客户名称", "申报金额", "申报币种", "流水付款方", "流水金额", "流水币种", "差额", "备注"]
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=h); cell.font = hfont; cell.fill = hf

    row_idx = 2
    for r in records:
        d_info = decl_map.get(r.declaration_id)
        c_info = cust_map.get(d_info.customer_id) if d_info else None
        # Apply search filter for export too
        if search and c_info:
            if search not in c_info.company_name: continue
        if search_code and c_info:
            if search_code not in (c_info.customer_code or ""): continue

        f_info = flow_map.get(r.flow_id)
        vals = [
            str(r.created_at)[:19] if r.created_at else "",
            c_info.customer_code if c_info else "",
            c_info.company_name if c_info else "",
            d_info.amount if d_info else "",
            d_info.currency if d_info else "",
            f_info.payer_name if f_info else "",
            f_info.amount if f_info else "",
            f_info.currency if f_info else "",
            r.amount_diff or 0,
            r.handling_note or "",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=c, value=str(v) if v is not None else "")
        row_idx += 1

    import io; output = io.BytesIO(); wb.save(output); output.seek(0)
    filename = f"reconciliation_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={filename}"})
